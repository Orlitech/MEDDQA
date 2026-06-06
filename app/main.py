"""
MedDQA System - Complete Main Application v3.0
Clinical Data Quality Assurance Platform
Multi-User | Real-Time | Professional
100% COMPLETE - All Features Included
"""

from fastapi import FastAPI, HTTPException, Request, Depends, Query
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pathlib import Path
from typing import Dict, Any, Optional, List
from sqlalchemy import text
from datetime import datetime, date, timedelta
from urllib.parse import unquote
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
import logging
import json
import secrets
import io
import os
import sys
import hashlib
from sqlalchemy import cast, Date

# ============================================================================
# PATH SETUP
# ============================================================================

def resource_path(relative_path):
    """Get absolute path to resource (works for PyInstaller)"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

STATIC_DIR = resource_path("app/static")
TEMPLATE_DIR = resource_path("app/templates")

# ============================================================================
# LOGGING
# ============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================================================
# FASTAPI APP
# ============================================================================

app = FastAPI(
    title="MedDQA",
    version="3.0.0",
    description="Clinical Data Quality Assurance Platform - Enhanced Edition"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

templates = Jinja2Templates(directory=TEMPLATE_DIR)

# ============================================================================
# IN-MEMORY STORES
# ============================================================================

active_sessions = {}  # user -> last_active_time
record_locks = {}     # hospital_number -> lock_info

# Authorized passkeys for regimen modification
AUTHORIZED_PASSKEYS = {
    "admin123": {"role": "admin", "name": "Administrator"},
    "dqa2024": {"role": "dqa_officer", "name": "DQA Officer"},
    "pharm2024": {"role": "pharmacist", "name": "Pharmacist"},
    "clinic2024": {"role": "clinician", "name": "Clinician"},
    "supervisor": {"role": "supervisor", "name": "Supervisor"},
    "1234": {"role": "admin", "name": "Administrator"}
}

# Session expiry time in minutes
SESSION_EXPIRY_MINUTES = 30

# ============================================================================
# SETUP CHECK
# ============================================================================

def is_setup_complete() -> bool:
    """Check if setup is complete"""
    env_file = Path('.env')
    if not env_file.exists():
        return False
    try:
        with open('.env', 'r') as f:
            content = f.read()
        # Check for default/demo values
        if 'EMR_DB_NAME=emr_database' in content and 'EMR_DB_PASSWORD=postgres' in content:
            return False
        # Check if required fields are filled
        required = ['EMR_DB_HOST', 'EMR_DB_NAME', 'EMR_DB_USER', 'EMR_DB_PASSWORD']
        for field in required:
            if f'{field}=' not in content:
                return False
        return True
    except Exception:
        return False

def get_database_connections():
    """Get database connections if configured"""
    try:
        from app.database import emr_engine, dqa_engine
        return emr_engine, dqa_engine
    except Exception as e:
        logger.warning(f"Database connections not available: {e}")
        return None, None

# ============================================================================
# MAIN PAGES
# ============================================================================

@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    """Main application page"""
    if not is_setup_complete():
        return RedirectResponse(url='/setup')
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/setup", response_class=HTMLResponse)
async def setup_page(request: Request):
    """Setup/configuration page"""
    if is_setup_complete():
        return RedirectResponse(url='/')
    return templates.TemplateResponse("setup.html", {"request": request})

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    emr_engine, dqa_engine = get_database_connections()
    
    return {
        "status": "healthy",
        "version": "3.0.0",
        "configured": is_setup_complete(),
        "emr_connected": emr_engine is not None,
        "dqa_connected": dqa_engine is not None,
        "active_users": len(active_sessions),
        "locked_records": len(record_locks),
        "timestamp": str(datetime.now())
    }

@app.get("/api/status")
async def api_status():
    """API status endpoint"""
    return {
        "success": True,
        "version": "3.0.0",
        "configured": is_setup_complete(),
        "timestamp": str(datetime.now())
    }

# ============================================================================
# SETUP API
# ============================================================================

@app.get("/api/setup/status")
async def get_setup_status():
    """Get current setup status"""
    return {
        "configured": is_setup_complete(),
        "env_exists": Path('.env').exists()
    }

@app.post("/api/setup/test-emr")
async def test_emr_connection(request: Request):
    """Test EMR database connection"""
    try:
        from app.database import test_emr_connection
        data = await request.json()
        result = test_emr_connection(
            host=data.get('host', 'localhost'),
            port=int(data.get('port', 5432)),
            dbname=data.get('dbname', ''),
            user=data.get('user', ''),
            password=data.get('password', '')
        )
        return result
    except Exception as e:
        logger.error(f"EMR test connection error: {e}")
        return {"success": False, "message": str(e)}

@app.post("/api/setup/test-dqa")
async def test_dqa_connection(request: Request):
    """Test DQA database connection"""
    try:
        from app.database import test_dqa_connection
        data = await request.json()
        result = test_dqa_connection(
            host=data.get('host', 'localhost'),
            port=int(data.get('port', 5432)),
            dbname=data.get('dbname', ''),
            user=data.get('user', ''),
            password=data.get('password', '')
        )
        return result
    except Exception as e:
        logger.error(f"DQA test connection error: {e}")
        return {"success": False, "message": str(e)}

@app.post("/api/setup/save")
async def save_setup(request: Request):
    """Save setup configuration"""
    try:
        data = await request.json()
        secret_key = secrets.token_urlsafe(32)
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # URL-encode passwords with special characters
        from urllib.parse import quote_plus
        
        emr_password = data.get('emr_password', '')
        dqa_password = data.get('dqa_password', '')
        
        # If password contains special chars, URL-encode it
        # But for .env files, we should use proper quoting
        # The best approach: wrap passwords in quotes in .env
        
        env_content = f"""# MedDQA Configuration - {timestamp}
APP_NAME=MedDQA
VERSION=3.0.0
DEBUG=True
HOST=0.0.0.0
PORT=8000

# EMR Database Configuration
EMR_DB_HOST={data.get('emr_host', 'localhost')}
EMR_DB_PORT={data.get('emr_port', 5432)}
EMR_DB_NAME={data.get('emr_dbname')}
EMR_DB_USER={data.get('emr_user')}
EMR_DB_PASSWORD='{emr_password}'

# DQA Database Configuration
DQA_DB_HOST={data.get('dqa_host', 'localhost')}
DQA_DB_PORT={data.get('dqa_port', 5432)}
DQA_DB_NAME={data.get('dqa_dbname', 'dqa_database')}
DQA_DB_USER={data.get('dqa_user')}
DQA_DB_PASSWORD='{dqa_password}'

# Security
SECRET_KEY={secret_key}
SESSION_EXPIRY_MINUTES=30
"""
        with open('.env', 'w') as f:
            f.write(env_content)
        
        logger.info("Configuration saved successfully")
        
        return {
            "success": True,
            "need_restart": True,
            "message": "Configuration saved! Please restart the application."
        }
    except Exception as e:
        logger.error(f"Save setup error: {e}")
        return {"success": False, "message": str(e)}

@app.post("/api/setup/reset")
async def reset_application():
    """Reset application configuration"""
    try:
        env_file = Path('.env')
        if env_file.exists():
            env_file.unlink()
        
        # Clear all in-memory data
        record_locks.clear()
        active_sessions.clear()
        
        logger.info("Application reset")
        
        return {
            "success": True,
            "message": "Application reset. Redirecting to setup..."
        }
    except Exception as e:
        logger.error(f"Reset error: {e}")
        return {"success": False, "message": str(e)}

@app.get("/api/setup/current-config")
async def get_current_config():
    """Get current configuration (masked)"""
    try:
        if not is_setup_complete():
            return {"configured": False}
        
        from app.config import settings
        
        return {
            "configured": True,
            "emr_host": settings.EMR_DB_HOST,
            "emr_port": settings.EMR_DB_PORT,
            "emr_dbname": settings.EMR_DB_NAME,
            "emr_user": settings.EMR_DB_USER,
            "dqa_dbname": settings.DQA_DB_NAME if hasattr(settings, 'DQA_DB_NAME') else 'N/A',
            "version": "3.0.0"
        }
    except Exception as e:
        logger.error(f"Get config error: {e}")
        return {"error": str(e)}

# ============================================================================
# ARV REFILL VALIDATION ENGINE
# ============================================================================

def parse_refill_date(date_str):
    """Parse date string to date object"""
    if not date_str:
        return None
    try:
        if isinstance(date_str, datetime):
            return date_str.date() if hasattr(date_str, 'date') else date_str
        return datetime.strptime(str(date_str)[:10], '%Y-%m-%d').date()
    except:
        return None

def add_days_to_date(date_obj, days):
    """Add days to a date"""
    if not date_obj:
        return None
    return date_obj + timedelta(days=int(days))

def days_difference(expected, actual):
    """Calculate days difference: negative=early, positive=late"""
    if not expected or not actual:
        return 0
    return (actual - expected).days

def format_date_str(date_obj):
    """Format date object to string"""
    if not date_obj:
        return 'N/A'
    return date_obj.strftime('%Y-%m-%d')

def get_sequence_label(idx):
    """Get human-readable sequence label"""
    if idx == 0: return "ART Initiation"
    elif idx == 1: return "First Refill"
    else: return f"Subsequent Refill #{idx}"

def validate_refill_timeline(refills):
    """
    Validate ARV refill timeline for a patient.
    ONLY validates ARV drugs - skips Anti-TB, Prophylaxis, and Other drugs.
    """
    if not refills or len(refills) == 0:
        return refills
    
    # Separate ARV drugs from non-ARV drugs
    arv_refills = []
    non_arv_refills = []
    
    for refill in refills:
        regimen_name = refill.get('regimen_name', '') or refill.get('regimen', '')
        drug_line = categorize_regimen(regimen_name)
        
        if drug_line == 'ARVs':
            arv_refills.append(refill)
        else:
            # Non-ARV drugs get marked as "Not Applicable"
            non_arv_refill = dict(refill)
            non_arv_refill['pickup_sequence'] = 0
            non_arv_refill['sequence_label'] = f"Non-ARV ({drug_line})"
            non_arv_refill['expected_next_date'] = 'N/A'
            non_arv_refill['days_early_or_late'] = 0
            non_arv_refill['refill_classification'] = f"{drug_line} Drug"
            non_arv_refill['validation_comment'] = f"Refill validation only applies to ARV drugs. This is a {drug_line} medication."
            non_arv_refill['fourteen_day_applies'] = False
            non_arv_refills.append(non_arv_refill)
    
    # Sort ARV refills by date (oldest first) for validation
    if arv_refills:
        sorted_arv = sorted(arv_refills, key=lambda r: r.get('pickup_date', ''))
        validated_arv = []
        
        for i, refill in enumerate(sorted_arv):
            pickup_date = parse_refill_date(refill.get('pickup_date'))
            duration = int(refill.get('duration', 0)) or 30
            
            if i == 0:
                # ART Initiation
                classification = "ART Initiation"
                comment = "First ARV pickup after enrollment. No validation required."
                days_diff = 0
                expected_date = pickup_date
                
            elif i == 1:
                # First Refill - NO 14-day allowance
                prev = sorted_arv[i - 1]
                prev_date = parse_refill_date(prev.get('pickup_date'))
                prev_dur = int(prev.get('duration', 0)) or 30
                expected_date = add_days_to_date(prev_date, prev_dur)
                days_diff = days_difference(expected_date, pickup_date) if pickup_date else 0
                
                if days_diff > 0:
                    classification = "Late Refill"
                    comment = f"First ARV refill {days_diff} day(s) late. Expected: {format_date_str(expected_date)} Need Tracking Form"
                elif days_diff >= -7:
                    classification = "On-Time Refill"
                    comment = f"First ARV refill {abs(days_diff)} day(s) early. Within acceptable range."
                else:
                    classification = "Excessively Early"
                    comment = f"First ARV refill {abs(days_diff)} day(s) early. First refill does NOT qualify for 14-day allowance."
                    
            else:
                # Second refill onward - 14-day allowance APPLIES
                prev = sorted_arv[i - 1]
                prev_date = parse_refill_date(prev.get('pickup_date'))
                prev_dur = int(prev.get('duration', 0)) or 30
                expected_date = add_days_to_date(prev_date, prev_dur)
                days_diff = days_difference(expected_date, pickup_date) if pickup_date else 0
                
                if days_diff > 7:
                    classification = "Late Refill"
                    comment = f"ARV refill {days_diff} day(s) late. Expected: {format_date_str(expected_date)}"
                elif days_diff >= -3:
                    classification = "On-Time Refill"
                    comment = "ARV refill on or near expected date."
                elif days_diff >= -14:
                    classification = "Acceptable Early"
                    comment = f"ARV picked up {abs(days_diff)} day(s) early. Within 14-day acceptable window."
                elif days_diff >= -30:
                    classification = "Excessively Early"
                    comment = f"ARV refill {abs(days_diff)} day(s) early. Exceeds 14-day allowance. Review required."
                else:
                    classification = "Possible Overlap"
                    comment = f"ARV refill {abs(days_diff)} day(s) early. Possible duplicate/overlapping refill. Urgent review required."
            
            # Add validation fields
            validated_refill = dict(refill)
            validated_refill['pickup_sequence'] = i + 1
            validated_refill['sequence_label'] = get_sequence_label(i)
            validated_refill['expected_next_date'] = format_date_str(expected_date) if expected_date else 'N/A'
            validated_refill['days_early_or_late'] = days_diff
            validated_refill['refill_classification'] = classification
            validated_refill['validation_comment'] = comment
            validated_refill['fourteen_day_applies'] = i >= 2
            
            validated_arv.append(validated_refill)
        
        # Combine validated ARV + non-ARV, sort newest first
        all_validated = validated_arv + non_arv_refills
        return sorted(all_validated, key=lambda r: r.get('pickup_date', ''), reverse=True)
    else:
        # No ARV drugs at all
        return sorted(non_arv_refills, key=lambda r: r.get('pickup_date', ''), reverse=True)
# ============================================================================
# PATIENT SEARCH API - FULL VERSION
# ============================================================================
@app.get("/api/patients/search/{hospital_number:path}")
async def search_patient(hospital_number: str, request: Request):
    """
    Search for a patient by hospital number.
    Uses the exact query structure that works with the EMR database.
    Extracts ALL drugs from JSONB regimens array.
    """
    try:
        from app.database import emr_engine
        
        if not emr_engine:
            return JSONResponse(
                status_code=500,
                content={"success": False, "detail": "Database not configured"}
            )
        
        hospital_number = unquote(hospital_number)
        user = request.headers.get("X-User", "anonymous")
        active_sessions[user] = datetime.now()
        
        logger.info(f"User '{user}' searching for: '{hospital_number}'")
        
        with emr_engine.connect() as conn:
            # Get patient with all fields
            result = conn.execute(
                text("""
                    SELECT 
                        p.uuid AS person_uuid,
                        p.hospital_number,
                        p.first_name,
                        p.surname,
                        p.other_name,
                        INITCAP(p.sex) AS sex,
                        p.date_of_birth,
                        p.date_of_registration AS date_enrolled,              
                        hac.visit_date::DATE AS art_start_date,                     
                        facility.name AS facility_name,
                        facility_lga.name AS lga,
                        facility_state.name AS state,
                        h.unique_id
                    FROM patient_person p
                    INNER JOIN hiv_enrollment h ON h.person_uuid = p.uuid AND h.archived = 0
                    INNER JOIN base_organisation_unit facility ON facility.id = h.facility_id
                    INNER JOIN base_organisation_unit facility_lga ON facility_lga.id = facility.parent_organisation_unit_id
                    INNER JOIN base_organisation_unit facility_state ON facility_state.id = facility_lga.parent_organisation_unit_id
                    LEFT JOIN hiv_art_clinical hac ON hac.hiv_enrollment_uuid = h.uuid 
                        AND hac.archived = 0 AND hac.is_commencement = TRUE
                    WHERE p.hospital_number = :hn AND p.archived = 0
                    LIMIT 1
                """),
                {"hn": hospital_number}
            )
            row = result.fetchone()
            
            if not row:
                # Try case-insensitive search
                result = conn.execute(
                    text("""
                        SELECT 
                            p.uuid AS person_uuid,
                            p.hospital_number,
                            p.first_name,
                            p.surname,
                            p.other_name,
                            INITCAP(p.sex) AS sex,
                            p.date_of_birth,
                            p.date_of_registration AS date_enrolled,
                            COALESCE(hac.visit_date::DATE, h.date_of_registration) AS art_start_date,
                            facility.name AS facility_name,
                            facility_lga.name AS lga,
                            facility_state.name AS state,
                            h.unique_id
                        FROM patient_person p
                        INNER JOIN hiv_enrollment h ON h.person_uuid = p.uuid AND h.archived = 0
                        INNER JOIN base_organisation_unit facility ON facility.id = h.facility_id
                        INNER JOIN base_organisation_unit facility_lga ON facility_lga.id = facility.parent_organisation_unit_id
                        INNER JOIN base_organisation_unit facility_state ON facility_state.id = facility_lga.parent_organisation_unit_id
                        LEFT JOIN hiv_art_clinical hac ON hac.hiv_enrollment_uuid = h.uuid 
                        AND hac.archived = 0 AND hac.is_commencement = TRUE
                        WHERE TRIM(p.hospital_number) ILIKE TRIM(:hn) AND p.archived = 0
                        LIMIT 1
                    """),
                    {"hn": hospital_number}
                )
                row = result.fetchone()
            
            if not row:
                return JSONResponse(
                    status_code=404,
                    content={"success": False, "detail": f"Patient not found: {hospital_number}"}
                )
            
            patient = dict(row._mapping)
            
            # Convert dates to strings for JSON serialization
            for key in ['date_of_birth', 'date_enrolled', 'art_start_date']:
                if patient.get(key):
                    patient[key] = str(patient[key])
            
            # If art_start_date is None, use date_enrolled as fallback
            if not patient.get('art_start_date'):
                patient['art_start_date'] = patient.get('date_enrolled')
            
            person_uuid = patient['person_uuid']
            logger.info(f"Patient found: {patient['first_name']} {patient['surname']} (UUID: {person_uuid})")
            
            # ================================================================
            # GET REFILLS - ALL DRUGS + HEIGHT/WEIGHT FROM triage_vital_sign
            # ================================================================
            refills = []
            try:
                refill_result = conn.execute(
                    text("""
                        SELECT 
                            hap.id AS visit_id,
                            hap.visit_date::DATE AS pickup_date,
                            hap.next_appointment,
                            hap.mmd_type,
                            COALESCE(
                                (elem.value ->> 'duration')::INTEGER,
                                (elem.value ->> 'prescribed')::INTEGER,
                                (elem.value ->> 'dispense')::INTEGER,
                                hap.refill_period,
                                0
                            ) AS duration,
                            COALESCE(
                                elem.value ->> 'regimenName',
                                elem.value ->> 'name',
                                'Unknown'
                            ) AS regimen_name,
                            (elem.value ->> 'regimenId')::BIGINT AS regimen_id,
                            COALESCE(hr.description, elem.value ->> 'regimenName', elem.value ->> 'name') AS regimen_full_name,
                            COALESCE(hrt.description, 'Other') AS regimen_line,
                            COALESCE(dsd.dsd_model, '') AS dsd_model,
                            -- Height and Weight from triage_vital_sign via hiv_art_clinical
                            tvs.body_weight AS weight_kg,
                            tvs.height AS height_cm,
                            -- Use visit_id + regimenId as unique identifier
                            hap.id::TEXT || '-' || (elem.value ->> 'regimenId') AS id,
                            elem.ordinality AS drug_index
                        FROM hiv_art_pharmacy hap
                        CROSS JOIN LATERAL jsonb_array_elements(hap.extra -> 'regimens') WITH ORDINALITY AS elem(value, ordinality)
                        LEFT JOIN hiv_regimen hr ON hr.id = (elem.value ->> 'regimenId')::BIGINT
                        LEFT JOIN hiv_regimen_type hrt ON hrt.id = hr.regimen_type_id
                        LEFT JOIN hiv_art_clinical hac ON hac.person_uuid = hap.person_uuid 
                            AND hac.visit_date = hap.visit_date 
                            AND hac.archived = 0
                        LEFT JOIN triage_vital_sign tvs ON tvs.uuid = hac.vital_sign_uuid 
                            AND tvs.archived = 0
                        LEFT JOIN (
                            SELECT DISTINCT ON (person_uuid) 
                                person_uuid, dsd_model
                            FROM dsd_devolvement WHERE archived = 0
                            ORDER BY person_uuid, date_devolved DESC
                        ) dsd ON dsd.person_uuid = hap.person_uuid
                        WHERE hap.person_uuid = :uuid
                        AND hap.archived = 0
                        AND hap.visit_date IS NOT NULL
                        AND hap.extra -> 'regimens' IS NOT NULL
                        AND jsonb_typeof(hap.extra -> 'regimens') = 'array'
                        AND jsonb_array_length(hap.extra -> 'regimens') > 0
                        ORDER BY hap.visit_date DESC, elem.ordinality
                    """),
                    {"uuid": person_uuid}
                )
                
                for r in refill_result:
                    refill = dict(r._mapping)
                    if refill.get('pickup_date'):
                        refill['pickup_date'] = str(refill['pickup_date'])
                    if refill.get('next_appointment'):
                        refill['next_appointment'] = str(refill['next_appointment'])
                    if not refill.get('regimen_name'):
                        refill['regimen_name'] = refill.get('regimen_full_name', 'Unknown')
                    refills.append(refill)
                    
                logger.info(f"Extracted {len(refills)} drug records with height/weight from triage_vital_sign")
                    
            except Exception as e:
                logger.warning(f"JSONB refill query failed: {e}")
                import traceback
                traceback.print_exc()
                try:
                    refill_result = conn.execute(
                        text("""
                            SELECT 
                                hap.id AS visit_id,
                                hap.visit_date AS pickup_date,
                                hap.next_appointment,
                                hap.mmd_type,
                                COALESCE(hap.refill_period, 0) AS duration,
                                'Unknown' AS regimen_name,
                                'Other' AS regimen_line,
                                COALESCE(dsd.dsd_model, '') AS dsd_model,
                                hap.id::TEXT AS id,
                                1 AS drug_index
                            FROM hiv_art_pharmacy hap
                            LEFT JOIN (
                                SELECT DISTINCT ON (person_uuid) 
                                    person_uuid,
                                    dsd_model
                                FROM dsd_devolvement
                                WHERE archived = 0
                                ORDER BY person_uuid, date_devolved DESC
                            ) dsd ON dsd.person_uuid = hap.person_uuid
                            WHERE hap.person_uuid = :uuid
                            AND hap.archived = 0
                            AND hap.visit_date IS NOT NULL
                            ORDER BY hap.visit_date DESC
                        """),
                        {"uuid": person_uuid}
                    )
                    for r in refill_result:
                        refill = dict(r._mapping)
                        if refill.get('pickup_date'):
                            refill['pickup_date'] = str(refill['pickup_date'])
                        if refill.get('next_appointment'):
                            refill['next_appointment'] = str(refill['next_appointment'])
                        refills.append(refill)
                    logger.info(f"Fallback: Found {len(refills)} refill records")
                except Exception as e2:
                    logger.error(f"Fallback refill query also failed: {e2}")
            
            # ================================================================
            # GET VIRAL LOADS
            # ================================================================
            viral_loads = []
            try:
                vl_result = conn.execute(
                    text("""
                        SELECT 
                            CAST(ls.date_sample_collected AS DATE) AS sample_collection_date,
                            sm.result_reported AS viral_load_result,
                            CAST(sm.date_result_reported AS DATE) AS result_date
                        FROM laboratory_result sm
                        INNER JOIN laboratory_test lt ON lt.id = sm.test_id
                        INNER JOIN laboratory_sample ls ON ls.test_id = lt.id
                        WHERE lt.lab_test_id = 16
                        AND sm.patient_uuid = :uuid
                        AND sm.result_reported IS NOT NULL
                        AND sm.archived = 0
                        ORDER BY ls.date_sample_collected DESC
                    """),
                    {"uuid": person_uuid}
                )
                for v in vl_result:
                    vl = dict(v._mapping)
                    if vl.get('sample_collection_date'):
                        vl['sample_collection_date'] = str(vl['sample_collection_date'])
                    if vl.get('result_date'):
                        vl['result_date'] = str(vl['result_date'])
                    viral_loads.append(vl)
                logger.info(f"Found {len(viral_loads)} viral load records")
            except Exception as e:
                logger.warning(f"VL query error: {e}")
            
            # ================================================================
            # GET CURRENT REGIMEN
            # ================================================================
            current_regimen = {}
            try:
                cr_result = conn.execute(
                    text("""
                        SELECT 
                            COALESCE(elem.value ->> 'regimenName', elem.value ->> 'name', 'Unknown') AS current_regimen,
                            COALESCE(hrt.description, 'Other') AS current_regimen_line,
                            hap.visit_date AS last_pickup_date,
                            hap.next_appointment
                        FROM hiv_art_pharmacy hap
                        CROSS JOIN LATERAL jsonb_array_elements(hap.extra -> 'regimens') WITH ORDINALITY AS elem(value, ordinality)
                        LEFT JOIN hiv_regimen hr ON hr.id = (elem.value ->> 'regimenId')::BIGINT
                        LEFT JOIN hiv_regimen_type hrt ON hrt.id = hr.regimen_type_id
                        WHERE hap.archived = 0 
                        AND hap.person_uuid = :uuid
                        AND hap.extra -> 'regimens' IS NOT NULL
                        AND jsonb_typeof(hap.extra -> 'regimens') = 'array'
                        ORDER BY hap.visit_date DESC
                        LIMIT 1
                    """),
                    {"uuid": person_uuid}
                )
                cr_row = cr_result.fetchone()
                if cr_row:
                    current_regimen = dict(cr_row._mapping)
                    if current_regimen.get('last_pickup_date'):
                        current_regimen['last_pickup_date'] = str(current_regimen['last_pickup_date'])
                    if current_regimen.get('next_appointment'):
                        current_regimen['next_appointment'] = str(current_regimen['next_appointment'])
            except Exception as e:
                logger.warning(f"Current regimen query error: {e}")
            
            # ================================================================
            # ✅ GET CLIENT VERIFICATION STATUS (ROC Verification)
            # ================================================================
            client_verification = None
            try:
                cv_result = conn.execute(
                    text("""
                        SELECT * FROM (
                            SELECT 
                                person_uuid, 
                                data->'attempt'->0->>'outcome' AS verification_outcome,
                                data->'attempt'->0->>'verificationStatus' AS verification_status,
                                CAST(data->'attempt'->0->>'dateOfAttempt' AS DATE) AS date_of_outcome,
                                ROW_NUMBER() OVER (PARTITION BY person_uuid ORDER BY CAST(data->'attempt'->0->>'dateOfAttempt' AS DATE) DESC)
                            FROM public.hiv_observation 
                            WHERE type = 'Client Verification' 
                            AND archived = 0 
                            AND CAST(data->'attempt'->0->>'dateOfAttempt' AS DATE) <= CURRENT_DATE 
                            AND CAST(data->'attempt'->0->>'dateOfAttempt' AS DATE) >= '1990-01-01'
                            AND person_uuid = :uuid
                        ) clientVerification 
                        WHERE row_number = 1 AND date_of_outcome IS NOT NULL
                    """),
                    {"uuid": person_uuid}
                )
                cv_row = cv_result.fetchone()
                if cv_row:
                    client_verification = {
                        "verification_outcome": cv_row[1] or 'N/A',
                        "verification_status": cv_row[2] or 'N/A',
                        "date_of_outcome": str(cv_row[3]) if cv_row[3] else 'N/A'
                    }
                    logger.info(f"Client verification found: {client_verification['verification_outcome']}")
            except Exception as e:
                logger.warning(f"Client verification query error: {e}")
            
            # ================================================================
            # GET CURRENT ART STATUS (from big query current_status logic)
            # ================================================================
            current_art_status = None
            try:
                status_result = conn.execute(
                    text("""
                        SELECT DISTINCT ON (pharmacy.person_uuid) 
                            pharmacy.person_uuid,
                            (CASE
                                WHEN stat.hiv_status ILIKE '%DEATH%' OR stat.hiv_status ILIKE '%Died%' THEN 'Died'
                                WHEN (stat.status_date > pharmacy.maxdate AND (stat.hiv_status ILIKE '%stop%' OR stat.hiv_status ILIKE '%out%' OR stat.hiv_status ILIKE '%Invalid %' OR stat.hiv_status ILIKE '%ART Transfer In%')) THEN stat.hiv_status
                                ELSE pharmacy.status
                            END) AS current_status,
                            (CASE
                                WHEN stat.hiv_status ILIKE '%DEATH%' OR stat.hiv_status ILIKE '%Died%' THEN stat.status_date
                                WHEN (stat.status_date > pharmacy.maxdate AND (stat.hiv_status ILIKE '%stop%' OR stat.hiv_status ILIKE '%out%' OR stat.hiv_status ILIKE '%Invalid %' OR stat.hiv_status ILIKE '%ART Transfer In%')) THEN stat.status_date
                                ELSE pharmacy.visit_date
                            END) AS status_date
                        FROM (
                            SELECT
                                (CASE
                                    WHEN hp.visit_date + hp.refill_period + INTERVAL '29 day' <= CURRENT_DATE THEN 'IIT'
                                    ELSE 'Active'
                                END) AS status,
                                (CASE
                                    WHEN hp.visit_date + hp.refill_period + INTERVAL '29 day' <= CURRENT_DATE THEN hp.visit_date + hp.refill_period + INTERVAL '29 day'
                                    ELSE hp.visit_date
                                END) AS visit_date,
                                hp.person_uuid,
                                MAXDATE
                            FROM hiv_art_pharmacy hp
                            INNER JOIN (
                                SELECT hap.person_uuid, hap.visit_date AS MAXDATE,
                                    ROW_NUMBER() OVER (PARTITION BY hap.person_uuid ORDER BY hap.visit_date DESC) AS rn
                                FROM hiv_art_pharmacy hap
                                INNER JOIN hiv_art_pharmacy_regimens pr ON pr.art_pharmacy_id = hap.id
                                INNER JOIN hiv_enrollment h ON h.person_uuid = hap.person_uuid AND h.archived = 0
                                INNER JOIN hiv_regimen r ON r.id = pr.regimens_id
                                INNER JOIN hiv_regimen_type rt ON rt.id = r.regimen_type_id
                                WHERE r.regimen_type_id IN (1,2,3,4,14,16)
                                AND hap.archived = 0
                                AND hap.visit_date <= CURRENT_DATE
                            ) MAX ON MAX.MAXDATE = hp.visit_date AND MAX.person_uuid = hp.person_uuid AND MAX.rn = 1
                            WHERE hp.archived = 0 AND hp.visit_date <= CURRENT_DATE
                        ) pharmacy
                        LEFT JOIN (
                            SELECT hst.hiv_status, hst.person_id, hst.cause_of_death,
                                hst.va_cause_of_death, hst.status_date
                            FROM (
                                SELECT * FROM (
                                    SELECT DISTINCT (person_id) AS person_id, status_date,
                                        cause_of_death, va_cause_of_death, hiv_status,
                                        ROW_NUMBER() OVER (PARTITION BY person_id ORDER BY status_date DESC)
                                    FROM hiv_status_tracker
                                    WHERE archived = 0 AND status_date <= CURRENT_DATE
                                ) s WHERE s.row_number = 1
                            ) hst
                            INNER JOIN hiv_enrollment he ON he.person_uuid = hst.person_id
                            WHERE hst.status_date <= CURRENT_DATE
                        ) stat ON stat.person_id = pharmacy.person_uuid
                        WHERE pharmacy.person_uuid = :uuid
                    """),
                    {"uuid": person_uuid}
                )
                status_row = status_result.fetchone()
                if status_row:
                    current_art_status = {
                        "status": status_row[1] or 'Unknown',
                        "status_date": str(status_row[2]) if status_row[2] else 'N/A'
                    }
                    logger.info(f"Current ART status: {current_art_status['status']}")
            except Exception as e:
                logger.warning(f"Current ART status query error: {e}")
            
            # ================================================================
            # RETURN ALL DATA
            # ================================================================
            lock_info = record_locks.get(hospital_number)
            
            # Apply refill validation
            if refills:
                refills = validate_refill_timeline(refills)
                logger.info(f"Validated {len(refills)} refills with timeline analysis")
            
            return {
                "success": True,
                "data": {
                    "patient_info": patient,
                    "refill_history": refills,
                    "viral_load_history": viral_loads,
                    "current_regimen": current_regimen,
                    "client_verification": client_verification,
                    "current_art_status": current_art_status
                },
                "lock_info": lock_info,
                "user": user
            }
            
    except Exception as e:
        logger.error(f"Search error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "detail": str(e)}
        )
# ============================================================================
# RECORD LOCKING
# ============================================================================

@app.post("/api/patients/{hospital_number:path}/lock")
async def lock_record(hospital_number: str, request: Request):
    """Lock a patient record for editing"""
    user = request.headers.get("X-User", "anonymous")
    hospital_number = unquote(hospital_number)
    
    # Check if already locked by another user
    if hospital_number in record_locks:
        existing = record_locks[hospital_number]
        if existing['user'] != user:
            # Check if lock has expired (30 minutes)
            lock_time = datetime.fromisoformat(existing['time'])
            if (datetime.now() - lock_time).seconds < 1800:
                return JSONResponse(
                    status_code=423,
                    content={
                        "success": False,
                        "detail": f"Record locked by {existing['user']} since {existing['time']}"
                    }
                )
    
    # Set/update lock
    record_locks[hospital_number] = {
        "user": user,
        "time": str(datetime.now()),
        "locked": True
    }
    
    logger.info(f"Record {hospital_number} locked by {user}")
    
    return {
        "success": True,
        "message": "Record locked",
        "locked_by": user,
        "locked_at": str(datetime.now())
    }

@app.post("/api/patients/{hospital_number:path}/unlock")
async def unlock_record(hospital_number: str, request: Request):
    """Unlock a patient record"""
    user = request.headers.get("X-User", "anonymous")
    hospital_number = unquote(hospital_number)
    
    if hospital_number in record_locks:
        if record_locks[hospital_number]['user'] == user:
            del record_locks[hospital_number]
            logger.info(f"Record {hospital_number} unlocked by {user}")
    
    return {"success": True, "message": "Record unlocked"}

# ============================================================================
# ART START DATE VALIDATION
# ============================================================================

@app.post("/api/patients/{hospital_number:path}/validate-art-start")
async def validate_art_start_date(hospital_number: str, request: Request):
    """
    Validate ART start date against first pickup date.
    If ART start date is after first pickup, flag as inconsistent.
    """
    try:
        from app.database import emr_engine
        
        data = await request.json()
        art_start_date = data.get('art_start_date')
        hospital_number = unquote(hospital_number)
        
        if not emr_engine:
            return JSONResponse(
                status_code=500,
                content={"success": False, "detail": "Database not configured"}
            )
        
        with emr_engine.connect() as conn:
            # Get patient UUID
            result = conn.execute(
                text("SELECT uuid FROM patient_person WHERE hospital_number = :hn AND archived = 0"),
                {"hn": hospital_number}
            )
            row = result.fetchone()
            
            if not row:
                return JSONResponse(
                    status_code=404,
                    content={"success": False, "detail": "Patient not found"}
                )
            
            person_uuid = row[0]
            
            # Get first pickup date (oldest visit_date)
            first_pickup = conn.execute(
                text("""
                    SELECT MIN(visit_date) AS first_pickup_date
                    FROM hiv_art_pharmacy
                    WHERE person_uuid = :uuid AND archived = 0 AND visit_date IS NOT NULL
                """),
                {"uuid": person_uuid}
            ).fetchone()
            
            first_pickup_date = str(first_pickup[0]) if first_pickup and first_pickup[0] else None
            
            if not first_pickup_date:
                return {
                    "success": True,
                    "is_consistent": True,
                    "message": "No pickups found to validate against",
                    "first_pickup_date": None,
                    "art_start_date": art_start_date
                }
            
            # Compare dates
            try:
                art_date = datetime.strptime(art_start_date[:10], '%Y-%m-%d').date()
                pickup_date = datetime.strptime(first_pickup_date[:10], '%Y-%m-%d').date()
                
                is_consistent = art_date <= pickup_date
                days_diff = abs((art_date - pickup_date).days)
                
                return {
                    "success": True,
                    "is_consistent": is_consistent,
                    "art_start_date": str(art_date),
                    "first_pickup_date": str(pickup_date),
                    "days_difference": days_diff,
                    "message": "✅ Dates are consistent" if is_consistent else
                               f"⚠️ ART Start Date is {days_diff} days AFTER first pickup"
                }
            except Exception as e:
                return {
                    "success": True,
                    "is_consistent": True,
                    "first_pickup_date": first_pickup_date,
                    "art_start_date": art_start_date,
                    "message": "Could not parse dates for comparison"
                }
                
    except Exception as e:
        logger.error(f"ART date validation error: {e}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "detail": str(e)}
        )

# ============================================================================
# PASSKEY VERIFICATION
# ============================================================================

@app.post("/api/auth/verify-passkey")
async def verify_passkey(request: Request):
    """Verify authorization passkey for sensitive operations"""
    try:
        data = await request.json()
        passkey = data.get('passkey', '')
        user = request.headers.get("X-User", "anonymous")
        
        if not passkey:
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "authorized": False,
                    "detail": "Passkey required"
                }
            )
        
        # Check passkey
        if passkey in AUTHORIZED_PASSKEYS:
            auth_info = AUTHORIZED_PASSKEYS[passkey]
            logger.info(f"User '{user}' authorized with passkey for role: {auth_info['role']}")
            
            return {
                "success": True,
                "authorized": True,
                "role": auth_info['role'],
                "authorized_by": auth_info['name'],
                "message": f"Authorized as {auth_info['name']} ({auth_info['role']})",
                "expires_in": "30 minutes"
            }
        
        # Log failed attempt
        logger.warning(f"Failed passkey attempt by user '{user}'")
        
        return JSONResponse(
            status_code=403,
            content={
                "success": False,
                "authorized": False,
                "detail": "Invalid passkey. Access denied."
            }
        )
        
    except Exception as e:
        logger.error(f"Passkey verification error: {e}")
        return JSONResponse(
            status_code=500,
            content={
                "success": False,
                "authorized": False,
                "detail": str(e)
            }
        )

@app.get("/api/auth/status")
async def auth_status(request: Request):
    """Check current authorization status"""
    user = request.headers.get("X-User", "anonymous")
    return {
        "user": user,
        "active_sessions": len(active_sessions),
        "available_roles": list(set(info['role'] for info in AUTHORIZED_PASSKEYS.values()))
    }

# ============================================================================
# EMR UPDATE - FULL JSONB SUPPORT - 100% WORKING
# ============================================================================

@app.put("/api/patients/update")
async def update_emr_record(request: Request):
    """
    Update EMR record fields.
    FULLY WORKING SUPPORT FOR:
    - ART Start Date
    - Sex (with gender JSONB)
    - Patient demographics (first_name, surname, other_name, date_of_birth)
    - Refill duration (updates both refill_period AND JSONB)
    - Refill regimen name (updates JSONB)
    - Height & Weight (in triage_vital_sign)
    - Visit date (syncs pharmacy, clinical, observation, triage)
    - Next appointment
    - Viral load results
    """
    try:
        from app.database import emr_engine
        
        if not emr_engine:
            return JSONResponse(
                status_code=500,
                content={"success": False, "detail": "Database not configured"}
            )
        
        data = await request.json()
        user = request.headers.get("X-User", "anonymous")
        
        hospital_number = data.get('hospital_number')
        field_name = data.get('field_name')
        new_value = data.get('new_value')
        record_type = data.get('record_type', 'patient')
        record_id = data.get('record_id')
        
        if not all([hospital_number, field_name, new_value is not None]):
            return JSONResponse(
                status_code=400,
                content={"success": False, "detail": "Missing required fields (hospital_number, field_name, new_value)"}
            )
        
        logger.info(f"User '{user}' updating {field_name} for {hospital_number} → {new_value}")
        
        with emr_engine.connect() as conn:
            trans = conn.begin()
            
            try:
                # Get patient UUID
                result = conn.execute(
                    text("SELECT uuid FROM patient_person WHERE hospital_number = :hn AND archived = 0"),
                    {"hn": hospital_number}
                )
                row = result.fetchone()
                
                if not row:
                    trans.rollback()
                    return JSONResponse(
                        status_code=404,
                        content={"success": False, "detail": "Patient not found"}
                    )
                
                person_uuid = row[0]
                
                # ============================================================
                # ART START DATE UPDATE
                # ============================================================
                if field_name == "art_start_date" and record_type == "patient":
                    conn.execute(
                        text("""
                            UPDATE hiv_art_clinical 
                            SET visit_date = :val 
                            WHERE hiv_enrollment_uuid = (
                                SELECT uuid FROM hiv_enrollment 
                                WHERE person_uuid = :uuid AND archived = 0
                            )
                            AND is_commencement = TRUE 
                            AND archived = 0
                        """),
                        {"val": new_value, "uuid": person_uuid}
                    )
                    trans.commit()
                    logger.info(f"✅ ART Start Date updated to {new_value}")
                    return {
                        "success": True,
                        "message": f"ART Start Date updated to {new_value}",
                        "updated_by": user,
                        "requires_validation": True
                    }
                
                # ============================================================
                # SEX UPDATE (with gender JSONB)
                # ============================================================
                elif field_name == "sex" and record_type == "patient":
                    valid_sexes = ['Male', 'Female', 'M', 'F', 'male', 'female', 'm', 'f']
                    if new_value not in valid_sexes:
                        trans.rollback()
                        return JSONResponse(
                            status_code=400,
                            content={
                                "success": False,
                                "detail": "Invalid sex value. Must be 'Male' or 'Female'"
                            }
                        )
                    
                    normalized_sex = 'Male' if new_value.lower() in ['male', 'm'] else 'Female'
                    gender_json = {"id": 376, "display": "Male"} if normalized_sex == 'Male' else {"id": 377, "display": "Female"}
                    
                    conn.execute(
                        text("""
                            UPDATE patient_person 
                            SET sex = :sex, gender = CAST(:gender AS jsonb) 
                            WHERE uuid = :uuid
                        """),
                        {"sex": normalized_sex, "gender": json.dumps(gender_json), "uuid": person_uuid}
                    )
                    
                    trans.commit()
                    return {
                        "success": True,
                        "message": f"Sex updated to {normalized_sex}",
                        "updated_by": user
                    }
                
                # ============================================================
                # PATIENT DEMOGRAPHICS FIELDS
                # ============================================================
                elif record_type == "patient":
                    allowed_fields = [
                        "first_name", "surname", "other_name",
                        "date_of_birth", "date_enrolled"
                    ]
                    
                    if field_name not in allowed_fields:
                        trans.rollback()
                        return JSONResponse(
                            status_code=400,
                            content={"success": False, "detail": f"Invalid patient field: {field_name}. Allowed: {', '.join(allowed_fields)}"}
                        )
                    
                    if field_name == "date_enrolled":
                        conn.execute(
                            text("""
                                UPDATE patient_person 
                                SET date_of_registration = :val 
                                WHERE uuid = :uuid
                            """),
                            {"val": new_value, "uuid": person_uuid}
                        )
                    else:
                        conn.execute(
                            text(f"UPDATE patient_person SET {field_name} = :val WHERE uuid = :uuid"),
                            {"val": new_value, "uuid": person_uuid}
                        )
                    
                    trans.commit()
                    return {
                        "success": True,
                        "message": f"{field_name} updated successfully",
                        "updated_by": user
                    }
                
                # ============================================================
                # REFILL UPDATE - TARGET BY regimenId
                # ============================================================
                elif record_type == "refill":
                    if not record_id:
                        trans.rollback()
                        return JSONResponse(status_code=400, content={
                            "success": False, "detail": "Record ID required"
                        })
                    
                    # Parse composite ID: "76921-116" → clean_id="76921", target_regimen_id="116"
                    clean_id = str(record_id)
                    target_regimen_id = None
                    
                    if '-' in str(record_id):
                        parts = str(record_id).split('-')
                        clean_id = parts[0]
                        try:
                            target_regimen_id = parts[1]
                        except (ValueError, IndexError):
                            target_regimen_id = None
                    
                    logger.info(f"🔧 Refill update: clean_id={clean_id}, target_regimen_id={target_regimen_id}, field={field_name}")
                    
                    # --------------------------------------------------------
                    # UPDATE DURATION - Target specific drug by regimenId
                    # --------------------------------------------------------
                    if field_name == "duration":
                        duration = int(new_value)
                        mmd = "MMD-1" if duration <= 30 else "MMD-2" if duration <= 60 else \
                              "MMD-3" if duration <= 90 else "MMD-4" if duration <= 120 else "MMD-6"
                        
                        # Update main refill_period
                        conn.execute(
                            text("UPDATE hiv_art_pharmacy SET refill_period = :val, mmd_type = :mmd WHERE id = :id"),
                            {"val": duration, "mmd": mmd, "id": clean_id}
                        )
                        
                        if target_regimen_id:
                            # Update ONLY the drug with matching regimenId
                            conn.execute(
                                text("""
                                    UPDATE hiv_art_pharmacy 
                                    SET extra = jsonb_set(
                                        extra,
                                        '{regimens}',
                                        (
                                            SELECT jsonb_agg(
                                                CASE 
                                                    WHEN (elem ->> 'regimenId') = :rid 
                                                    THEN elem || jsonb_build_object(
                                                        'duration', :d,
                                                        'prescribed', :p,
                                                        'dispense', :disp
                                                    )
                                                    ELSE elem
                                                END
                                            )
                                            FROM jsonb_array_elements(extra -> 'regimens') AS elem
                                        )
                                    )
                                    WHERE id = :id
                                    AND extra -> 'regimens' IS NOT NULL
                                    AND jsonb_typeof(extra -> 'regimens') = 'array'
                                """),
                                {
                                    "d": duration,
                                    "p": duration,
                                    "disp": str(duration),
                                    "id": clean_id,
                                    "rid": target_regimen_id
                                }
                            )
                        else:
                            # Fallback: update all drugs
                            conn.execute(
                                text("""
                                    UPDATE hiv_art_pharmacy 
                                    SET extra = jsonb_set(
                                        extra,
                                        '{regimens}',
                                        (
                                            SELECT jsonb_agg(
                                                elem || jsonb_build_object(
                                                    'duration', :d,
                                                    'prescribed', :p,
                                                    'dispense', :disp
                                                )
                                            )
                                            FROM jsonb_array_elements(extra -> 'regimens') AS elem
                                        )
                                    )
                                    WHERE id = :id
                                """),
                                {"d": duration, "p": duration, "disp": str(duration), "id": clean_id}
                            )
                        
                        logger.info(f"✅ Duration updated to {duration} days for regimen_id={target_regimen_id}")
                    
                    # --------------------------------------------------------
                    # UPDATE REGIMEN NAME - Target specific drug by regimenId
                    # --------------------------------------------------------
                    elif field_name == "regimen":
                        # ✅ Convert to proper JSON string first
                        regimen_json = json.dumps(str(new_value))  # This wraps in quotes: "TAF(25mg)..."
                        
                        if target_regimen_id:
                            conn.execute(
                                text("""
                                    UPDATE hiv_art_pharmacy 
                                    SET extra = jsonb_set(
                                        extra,
                                        '{regimens}',
                                        (
                                            SELECT jsonb_agg(
                                                CASE 
                                                    WHEN (elem ->> 'regimenId') = :rid 
                                                    THEN elem || jsonb_build_object('regimenName', CAST(:rn AS jsonb))
                                                    ELSE elem
                                                END
                                            )
                                            FROM jsonb_array_elements(extra -> 'regimens') AS elem
                                        )
                                    )
                                    WHERE id = :id
                                    AND extra -> 'regimens' IS NOT NULL
                                    AND jsonb_typeof(extra -> 'regimens') = 'array'
                                """),
                                {"rn": regimen_json, "id": clean_id, "rid": str(target_regimen_id)}
                            )
                        else:
                            conn.execute(
                                text("""
                                    UPDATE hiv_art_pharmacy 
                                    SET extra = jsonb_set(
                                        extra,
                                        '{regimens}',
                                        (
                                            SELECT jsonb_agg(
                                                elem || jsonb_build_object('regimenName', CAST(:rn AS jsonb))
                                            )
                                            FROM jsonb_array_elements(extra -> 'regimens') AS elem
                                        )
                                    )
                                    WHERE id = :id
                                """),
                                {"rn": regimen_json, "id": clean_id}
                            )
                        
                        logger.info(f"✅ Regimen updated to '{new_value}' for regimen_id={target_regimen_id}")
                    
                    # --------------------------------------------------------
                    # UPDATE HEIGHT (in triage_vital_sign via hiv_art_clinical)
                    # --------------------------------------------------------
                    elif field_name == "height":
                        conn.execute(
                            text("""
                                UPDATE triage_vital_sign 
                                SET height = CAST(:val AS NUMERIC)
                                WHERE uuid IN (
                                    SELECT hac.vital_sign_uuid 
                                    FROM hiv_art_clinical hac 
                                    WHERE hac.person_uuid = (
                                        SELECT person_uuid FROM hiv_art_pharmacy WHERE id = :pharmacy_id
                                    )
                                    AND hac.visit_date = (
                                        SELECT visit_date FROM hiv_art_pharmacy WHERE id = :pharmacy_id2
                                    )
                                    AND hac.archived = 0
                                )
                                AND archived = 0
                            """),
                            {"val": float(new_value), "pharmacy_id": clean_id, "pharmacy_id2": clean_id}
                        )
                        logger.info(f"✅ Height updated to {new_value} cm for visit {clean_id}")
                    
                    # --------------------------------------------------------
                    # UPDATE WEIGHT (in triage_vital_sign via hiv_art_clinical)
                    # --------------------------------------------------------
                    elif field_name == "weight":
                        conn.execute(
                            text("""
                                UPDATE triage_vital_sign 
                                SET body_weight = CAST(:val AS NUMERIC)
                                WHERE uuid IN (
                                    SELECT hac.vital_sign_uuid 
                                    FROM hiv_art_clinical hac 
                                    WHERE hac.person_uuid = (
                                        SELECT person_uuid FROM hiv_art_pharmacy WHERE id = :pharmacy_id
                                    )
                                    AND hac.visit_date = (
                                        SELECT visit_date FROM hiv_art_pharmacy WHERE id = :pharmacy_id2
                                    )
                                    AND hac.archived = 0
                                )
                                AND archived = 0
                            """),
                            {"val": float(new_value), "pharmacy_id": clean_id, "pharmacy_id2": clean_id}
                        )
                        logger.info(f"✅ Weight updated to {new_value} kg for visit {clean_id}")
                    
                    # --------------------------------------------------------
                    # ✅ UPDATE VISIT DATE - Sync ALL related tables
                    # --------------------------------------------------------
                    elif field_name == "visit_date":
                        # Get OLD date first
                        old_date_result = conn.execute(
                            text("SELECT visit_date FROM hiv_art_pharmacy WHERE id = :id"),
                            {"id": clean_id}
                        )
                        old_date_row = old_date_result.fetchone()
                        old_date = old_date_row[0] if old_date_row else None
                        
                        # 1. Update hiv_art_pharmacy (the pharmacy visit)
                        conn.execute(
                            text("UPDATE hiv_art_pharmacy SET visit_date = :val WHERE id = :id"),
                            {"val": new_value, "id": clean_id}
                        )
                        logger.info(f"✅ Pharmacy visit_date updated to {new_value}")
                        
                        if old_date:
                            # 2. Update hiv_art_clinical (clinical data for this visit)
                            conn.execute(
                                text("""
                                    UPDATE hiv_art_clinical 
                                    SET visit_date = CAST(:val AS DATE)
                                    WHERE person_uuid = (
                                        SELECT person_uuid FROM hiv_art_pharmacy WHERE id = :pharmacy_id
                                    )
                                    AND visit_date = CAST(:old_date AS DATE)
                                    AND archived = 0
                                """),
                                {"val": new_value, "pharmacy_id": clean_id, "old_date": str(old_date)}
                            )
                            logger.info(f"✅ hiv_art_clinical date synced to {new_value}")
                            
                            # 3. Update hiv_observation (observations for this date)
                            conn.execute(
                                text("""
                                    UPDATE hiv_observation 
                                    SET date_of_observation = CAST(:val AS DATE)
                                    WHERE person_uuid = (
                                        SELECT person_uuid FROM hiv_art_pharmacy WHERE id = :pharmacy_id
                                    )
                                    AND date_of_observation = CAST(:old_date AS DATE)
                                    AND archived = 0
                                """),
                                {"val": new_value, "pharmacy_id": clean_id, "old_date": str(old_date)}
                            )
                            logger.info(f"✅ hiv_observation date synced to {new_value}")
                            
                            # 4. triage_vital_sign is linked by UUID (not date), so it follows clinical record
                            logger.info(f"✅ All related records synced from {old_date} to {new_value}")
                    
                    # --------------------------------------------------------
                    # UPDATE NEXT APPOINTMENT
                    # --------------------------------------------------------
                    elif field_name == "next_appointment":
                        conn.execute(
                            text("UPDATE hiv_art_pharmacy SET next_appointment = :val WHERE id = :id"),
                            {"val": new_value, "id": clean_id}
                        )
                        logger.info(f"✅ next_appointment updated to {new_value}")
                    
                    # --------------------------------------------------------
                    # UPDATE MMD TYPE
                    # --------------------------------------------------------
                    elif field_name == "mmd_type":
                        conn.execute(
                            text("UPDATE hiv_art_pharmacy SET mmd_type = :val WHERE id = :id"),
                            {"val": new_value, "id": clean_id}
                        )
                        logger.info(f"✅ mmd_type updated to {new_value}")
                    
                    else:
                        trans.rollback()
                        return JSONResponse(status_code=400, content={
                            "success": False,
                            "detail": f"Invalid refill field: {field_name}. Valid fields: duration, regimen, height, weight, visit_date, next_appointment, mmd_type"
                        })
                    
                    trans.commit()
                    return {
                        "success": True,
                        "message": f"{field_name} updated successfully",
                        "updated_by": user,
                        "refill_id": clean_id
                    }
                
                # ============================================================
                # VIRAL LOAD UPDATE
                # ============================================================
                elif record_type == "viral_load":
                    trans.rollback()

                    return JSONResponse(
                        status_code=400,
                        content={
                            "success": False,
                            "detail": "Use /api/patients/update-vl for Viral Load updates"
                        }
                    )
                    
                else:
                    trans.rollback()
                    return JSONResponse(status_code=400, content={
                        "success": False, "detail": f"Unknown record type: {record_type}"
                    })
                    
            except Exception as e:
                trans.rollback()
                raise e
                
    except Exception as e:
        logger.error(f"Update error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})
# ============================================================================
# REGIMEN LIST
# ============================================================================

@app.get("/api/reference/regimens")
async def get_regimens():
    """Get list of available regimens grouped by clinical category"""
    try:
        from app.database import emr_engine
        
        if emr_engine:
            with emr_engine.connect() as conn:
                result = conn.execute(text("""
                    SELECT DISTINCT 
                        hr.description AS regimen_name, 
                        hrt.description AS regimen_line
                    FROM hiv_regimen hr
                    INNER JOIN hiv_regimen_type hrt ON hrt.id = hr.regimen_type_id
                    ORDER BY hrt.description, hr.description
                """))
                
                regimens = []
                for row in result:
                    name = row[0]
                    original_line = row[1] if row[1] else 'Other'
                    
                    # Re-categorize based on drug name
                    line = categorize_regimen(name)
                    
                    regimens.append({"name": name, "line": line})
                
                # Remove duplicates (same name, same line)
                seen = set()
                unique_regimens = []
                for r in regimens:
                    key = (r['name'], r['line'])
                    if key not in seen:
                        seen.add(key)
                        unique_regimens.append(r)
                
                logger.info(f"Loaded {len(unique_regimens)} regimens from database")
                
                return {
                    "success": True,
                    "count": len(unique_regimens),
                    "regimens": unique_regimens
                }
        
        # Fallback regimens
        return get_fallback_regimens()
            
    except Exception as e:
        logger.error(f"Regimen list error: {e}")
        return get_fallback_regimens()

def get_fallback_regimens():
    """Fallback regimens when database is not available"""
    return {
        "success": True,
        "count": 17,
        "regimens": [
            # ARVs
            {"name": "TDF/3TC/DTG", "line": "ARVs"},
            {"name": "ABC/3TC/DTG", "line": "ARVs"},
            {"name": "AZT/3TC/DTG", "line": "ARVs"},
            {"name": "AZT/3TC/EFV", "line": "ARVs"},
            {"name": "AZT/3TC/NVP", "line": "ARVs"},
            {"name": "AZT/3TC/LPV/r", "line": "ARVs"},
           
            # Anti-TB
            {"name": "Isoniazid (INH) 300mg", "line": "Anti-TB"},
            {"name": "3HP (Isoniazid + Rifapentine)", "line": "Anti-TB"},
            # Prophylaxis
            {"name": "Cotrimoxazole 960mg", "line": "Prophylaxis"},
            {"name": "Cotrimoxazole 480mg", "line": "Prophylaxis"},
            {"name": "Fluconazole 200mg", "line": "Prophylaxis"},
            # Other
            {"name": "Pyridoxine 50mg", "line": "Other"}
        ]
    }
# ============================================================================
# CARE CARD SAVE & LOAD
# ============================================================================
@app.post("/api/care-cards/save")
async def save_care_card_data(request: Request):
    """Save care card data for a patient"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import CareCardRecord
        from sqlalchemy.orm import Session
        
        data = await request.json()
        user = request.headers.get("X-User", "anonymous")
        
        hospital_number = data.get('hospital_number')
        drug_pickups = data.get('drug_pickups', [])
        viral_loads = data.get('viral_loads', [])
        person_uuid = data.get('person_uuid', '')
        
        if not hospital_number:
            return JSONResponse(
                status_code=400,
                content={"success": False, "detail": "Hospital number required"}
            )
        
        with Session(dqa_engine) as dqa_session:
            existing = dqa_session.query(CareCardRecord).filter(
                CareCardRecord.hospital_number == hospital_number
            ).first()
            
            if existing:
                existing.drug_pickups = drug_pickups
                existing.viral_loads = viral_loads
                existing.updated_by = user
                existing.updated_at = datetime.now()
                action = "updated"
            else:
                new_record = CareCardRecord(
                    hospital_number=hospital_number,
                    person_uuid=person_uuid,
                    drug_pickups=drug_pickups,
                    viral_loads=viral_loads,
                    created_by=user,
                    updated_by=user
                )
                dqa_session.add(new_record)
                action = "created"
            
            dqa_session.commit()
            
            logger.info(f"Care card {action} for {hospital_number} by {user}")
            
            return {
                "success": True,
                "action": action,
                "message": f"Care card data {action} successfully",
                "updated_by": user
            }
            
    except Exception as e:
        logger.error(f"Save care card error: {e}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "detail": str(e)}
        )

@app.get("/api/care-cards/load/{hospital_number:path}")
async def load_care_card_data(hospital_number: str, request: Request):
    """Load previously saved care card data"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import CareCardRecord
        from sqlalchemy.orm import Session
        
        hospital_number = unquote(hospital_number)
        user = request.headers.get("X-User", "anonymous")
        
        with Session(dqa_engine) as dqa_session:
            record = dqa_session.query(CareCardRecord).filter(
                CareCardRecord.hospital_number == hospital_number
            ).first()
            
            if record:
                return {
                    "success": True,
                    "found": True,
                    "data": {
                        "hospital_number": record.hospital_number,
                        "drug_pickups": record.drug_pickups or [],
                        "viral_loads": record.viral_loads or [],
                        "enrollment_data": record.enrollment_data or {},
                        "created_by": record.created_by,
                        "updated_by": record.updated_by,
                        "created_at": str(record.created_at) if record.created_at else None,
                        "updated_at": str(record.updated_at) if record.updated_at else None,
                        "is_verified": record.is_verified,
                        "verified_by": record.verified_by,
                        "verified_at": str(record.verified_at) if record.verified_at else None
                    }
                }
            else:
                return {"success": True, "found": False, "data": None}
                
    except Exception as e:
        logger.error(f"Load care card error: {e}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "detail": str(e)}
        )
    
# ============================================================================
# CARE CARD COMPARISON
# ============================================================================

@app.post("/api/care-cards/compare")
async def compare_care_card(request: Request):
    """Compare care card entries with EMR records"""
    try:
        from app.database import emr_engine
        
        data = await request.json()
        hospital_number = data.get('hospital_number')
        user = request.headers.get("X-User", "anonymous")
        
        if not hospital_number:
            return JSONResponse(
                status_code=400,
                content={"success": False, "detail": "Hospital number required"}
            )
        
        logger.info(f"User '{user}' comparing care card for: {hospital_number}")
        
        with emr_engine.connect() as conn:
            # Get patient
            result = conn.execute(
                text("""
                    SELECT p.uuid AS person_uuid 
                    FROM patient_person p
                    WHERE p.hospital_number = :hn AND p.archived = 0 
                    LIMIT 1
                """),
                {"hn": hospital_number}
            )
            patient_row = result.fetchone()
            
            if not patient_row:
                return JSONResponse(
                    status_code=404,
                    content={"success": False, "detail": "Patient not found"}
                )
            
            person_uuid = patient_row[0]
            
            # Get EMR refills
            emr_refills = []
            refill_result = conn.execute(
                text("""
                    SELECT 
                        visit_date AS pickup_date, 
                        COALESCE(refill_period, 0) AS duration,
                        next_appointment
                    FROM hiv_art_pharmacy 
                    WHERE person_uuid = :uuid AND archived = 0
                    ORDER BY visit_date DESC
                """),
                {"uuid": person_uuid}
            )
            for r in refill_result:
                refill = dict(r._mapping)
                if refill.get('pickup_date'):
                    refill['pickup_date'] = str(refill['pickup_date'])
                if refill.get('next_appointment'):
                    refill['next_appointment'] = str(refill['next_appointment'])
                emr_refills.append(refill)
            
            # Get EMR viral loads
            emr_vls = []
            vl_result = conn.execute(
                text("""
                    SELECT 
                        CAST(ls.date_sample_collected AS DATE) AS sample_collection_date,
                        sm.result_reported AS viral_load_result,
                        CAST(sm.date_result_reported AS DATE) AS result_date
                    FROM laboratory_result sm
                    INNER JOIN laboratory_test lt ON lt.id = sm.test_id
                    INNER JOIN laboratory_sample ls ON ls.test_id = lt.id
                    WHERE lt.lab_test_id = 16 
                    AND sm.patient_uuid = :uuid
                    AND sm.result_reported IS NOT NULL 
                    AND sm.archived = 0
                    ORDER BY ls.date_sample_collected DESC
                """),
                {"uuid": person_uuid}
            )
            for v in vl_result:
                vl = dict(v._mapping)
                if vl.get('sample_collection_date'):
                    vl['sample_collection_date'] = str(vl['sample_collection_date'])
                if vl.get('result_date'):
                    vl['result_date'] = str(vl['result_date'])
                emr_vls.append(vl)
        
        # Perform comparison
        comparison_results = []
        care_pickups = data.get('drug_pickups', [])
        
        for i, care_pickup in enumerate(care_pickups):
            emr_pickup = emr_refills[i] if i < len(emr_refills) else None
            
            if not emr_pickup:
                comparison_results.append({
                    "field_name": f"Pickup #{i+1}",
                    "emr_value": "Not in EMR",
                    "care_card_value": str(care_pickup.get('pickup_date', 'N/A')),
                    "match": False
                })
                continue
            
            # Compare date
            emr_date = str(emr_pickup.get('pickup_date', ''))[:10]
            cc_date = str(care_pickup.get('pickup_date', ''))[:10]
            comparison_results.append({
                "field_name": f"Pickup #{i+1} Date",
                "emr_value": emr_date or 'N/A',
                "care_card_value": cc_date or 'N/A',
                "match": emr_date == cc_date
            })
            
            # Compare duration
            emr_dur = str(emr_pickup.get('duration', 0))
            cc_dur = str(care_pickup.get('duration', 0))
            comparison_results.append({
                "field_name": f"Pickup #{i+1} Duration",
                "emr_value": f"{emr_dur} days",
                "care_card_value": f"{cc_dur} days",
                "match": emr_dur == cc_dur
            })
        
        # Compare viral loads
        care_vls = data.get('viral_loads', [])
        for i, care_vl in enumerate(care_vls):
            emr_vl = emr_vls[i] if i < len(emr_vls) else None
            
            if not emr_vl:
                comparison_results.append({
                    "field_name": f"VL #{i+1}",
                    "emr_value": "Not in EMR",
                    "care_card_value": str(care_vl.get('viral_load_result', 'N/A')),
                    "match": False
                })
                continue
            
            emr_date = str(emr_vl.get('sample_collection_date', ''))[:10]
            cc_date = str(care_vl.get('sample_collection_date', ''))[:10]
            comparison_results.append({
                "field_name": f"VL #{i+1} Sample Date",
                "emr_value": emr_date or 'N/A',
                "care_card_value": cc_date or 'N/A',
                "match": emr_date == cc_date
            })
            
            emr_result = str(emr_vl.get('viral_load_result', '')).strip()
            cc_result = str(care_vl.get('viral_load_result', '')).strip()
            comparison_results.append({
                "field_name": f"VL #{i+1} Result",
                "emr_value": emr_result or 'N/A',
                "care_card_value": cc_result or 'N/A',
                "match": emr_result.lower() == cc_result.lower()
            })
        
        all_matched = all(r['match'] for r in comparison_results)
        mismatch_count = sum(1 for r in comparison_results if not r['match'])
        
        return {
            "success": True,
            "hospital_number": hospital_number,
            "comparison_results": comparison_results,
            "all_matched": all_matched,
            "can_submit": all_matched,
            "mismatch_count": mismatch_count,
            "matched_count": len(comparison_results) - mismatch_count,
            "message": "✅ All records match!" if all_matched else f"❌ Found {mismatch_count} discrepancies"
        }
        
    except Exception as e:
        logger.error(f"Comparison error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "detail": str(e)}
        )
# ============================================================================
# DQA SUBMISSION
# ============================================================================

@app.post("/api/care-cards/submit")
async def submit_care_card(request: Request):
    """Submit DQA verification data with full correction details"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import DQAAuditLog, CareCardRecord
        from sqlalchemy.orm import Session
        
        data = await request.json()
        user = request.headers.get("X-User", "anonymous")
        
        hospital_number = data.get('hospital_number')
        care_card_data = data.get('care_card_data', {})
        comparison_results = data.get('comparison_results', [])
        total_comparisons = data.get('total_comparisons', len(comparison_results))
        matched_comparisons = data.get('matched_comparisons', sum(1 for r in comparison_results if r.get('match')))
        
        if not hospital_number:
            return JSONResponse(
                status_code=400,
                content={"success": False, "detail": "Hospital number required"}
            )
        
        logger.info(f"User '{user}' submitting DQA for: {hospital_number}")
        
        with Session(dqa_engine) as dqa_session:
            # Check if already verified
            existing_record = dqa_session.query(CareCardRecord).filter(
                CareCardRecord.hospital_number == hospital_number,
                CareCardRecord.is_verified == True
            ).first()
            
            if existing_record:
                return {
                    "success": True,
                    "message": f"This patient was already verified by {existing_record.verified_by}",
                    "already_verified": True,
                    "verified_by": existing_record.verified_by,
                    "verified_at": str(existing_record.verified_at) if existing_record.verified_at else None,
                    "submitted_by": user,
                    "timestamp": str(datetime.now())
                }
            
            # Create audit log
            audit_log = DQAAuditLog(
                hospital_number=hospital_number,
                person_uuid=data.get('person_uuid', ''),
                first_name=data.get('first_name', ''),
                surname=data.get('surname', ''),
                facility_name=data.get('facility_name', ''),
                state=data.get('state', ''),
                care_card_data=care_card_data,
                emr_snapshot={},
                validation_status='Matched' if matched_comparisons == total_comparisons else 'Partial Match',
                discrepancies_found=total_comparisons - matched_comparisons,
                issues_fixed=0,
                total_comparisons=total_comparisons,
                matched_comparisons=matched_comparisons,
                user_name=user
            )
            dqa_session.add(audit_log)
            
            # Update care card record
            care_record = dqa_session.query(CareCardRecord).filter(
                CareCardRecord.hospital_number == hospital_number
            ).first()
            
            if care_record:
                care_record.is_verified = True
                care_record.verified_by = user
                care_record.verified_at = datetime.now()
            
            dqa_session.commit()
            
            logger.info(f"DQA submitted for {hospital_number} by {user}, Audit ID: {audit_log.id}")
            
            return {
                "success": True,
                "message": "DQA verification submitted successfully!",
                "submitted_by": user,
                "audit_id": audit_log.id,
                "already_verified": False,
                "timestamp": str(datetime.now())
            }
            
    except Exception as e:
        logger.error(f"Submission error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "detail": str(e)}
        )

# ============================================================================
# DQA REPORTS
# ============================================================================

@app.get("/api/reports/summary")
async def get_dqa_summary(request: Request):
    """Get DQA summary statistics"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import DQAAuditLog, CareCardRecord
        from sqlalchemy.orm import Session
        from sqlalchemy import func
        
        with Session(dqa_engine) as session:
            total_audits = session.query(func.count(DQAAuditLog.id)).scalar() or 0
            total_verified = session.query(func.count(CareCardRecord.id)).filter(
                CareCardRecord.is_verified == True
            ).scalar() or 0
            total_care_cards = session.query(func.count(CareCardRecord.id)).scalar() or 0
            
            return {
                "success": True,
                "summary": {
                    "total_audits": total_audits,
                    "total_verified": total_verified,
                    "total_care_cards": total_care_cards,
                    "verification_rate": f"{(total_verified/total_care_cards*100):.1f}%" if total_care_cards > 0 else "0%"
                }
            }
    except Exception as e:
        logger.error(f"Summary error: {e}")
        return {"success": False, "detail": str(e)}

@app.get("/api/reports/excel")
async def generate_excel_report(request: Request):
    """Generate Excel report of DQA data"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import CareCardRecord, DQAAuditLog
        from sqlalchemy.orm import Session
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        user = request.headers.get("X-User", "anonymous")
        logger.info(f"User '{user}' generating DQA report")
        
        wb = Workbook()
        
        # === Sheet 1: Summary ===
        ws_summary = wb.active
        ws_summary.title = "DQA Summary"
        
        # Title
        ws_summary.merge_cells('A1:G1')
        ws_summary['A1'] = "MedDQA - Data Quality Audit Report"
        ws_summary['A1'].font = Font(bold=True, size=16, color="1e40af")
        ws_summary['A1'].alignment = Alignment(horizontal="center")
        
        ws_summary.merge_cells('A2:G2')
        ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | By: {user}"
        ws_summary['A2'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = [
            "Hospital Number", "Patient Name", "Drug Pickups",
            "Viral Loads", "Verified", "Verified By", "Verified Date"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        with Session(dqa_engine) as session:
            records = session.query(CareCardRecord).order_by(
                CareCardRecord.created_at.desc()
            ).all()
            
            for row_idx, record in enumerate(records, 5):
                ws_summary.cell(row=row_idx, column=1, value=record.hospital_number)
                ws_summary.cell(row=row_idx, column=2, value=record.created_by or '')
                ws_summary.cell(row=row_idx, column=3, value=len(record.drug_pickups or []))
                ws_summary.cell(row=row_idx, column=4, value=len(record.viral_loads or []))
                ws_summary.cell(row=row_idx, column=5, value="Yes" if record.is_verified else "No")
                ws_summary.cell(row=row_idx, column=6, value=record.verified_by or '')
                ws_summary.cell(row=row_idx, column=7, value=str(record.verified_at)[:19] if record.verified_at else '')
        
        # Auto-size columns
        for column in ws_summary.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_summary.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # === Sheet 2: Audit Log ===
        ws_audit = wb.create_sheet("Audit Log")
        
        audit_headers = [
            "ID", "Hospital Number", "User", "Status",
            "Discrepancies", "Comparisons", "Matched", "Timestamp"
        ]
        
        for col, header in enumerate(audit_headers, 1):
            cell = ws_audit.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        with Session(dqa_engine) as session:
            audits = session.query(DQAAuditLog).order_by(
                DQAAuditLog.created_at.desc()
            ).limit(500).all()
            
            for row_idx, audit in enumerate(audits, 2):
                ws_audit.cell(row=row_idx, column=1, value=audit.id)
                ws_audit.cell(row=row_idx, column=2, value=audit.hospital_number)
                ws_audit.cell(row=row_idx, column=3, value=audit.user_name)
                ws_audit.cell(row=row_idx, column=4, value=audit.validation_status)
                ws_audit.cell(row=row_idx, column=5, value=audit.discrepancies_found)
                ws_audit.cell(row=row_idx, column=6, value=audit.total_comparisons)
                ws_audit.cell(row=row_idx, column=7, value=audit.matched_comparisons)
                ws_audit.cell(row=row_idx, column=8, value=str(audit.created_at)[:19] if audit.created_at else '')
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"DQA_Report_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Report error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "detail": str(e)}
        )

# ============================================================================
# TEAM / ACTIVE USERS
# ============================================================================

@app.get("/api/team/active-users")
async def get_active_users():
    """Get list of currently active users"""
    # Clean up expired sessions
    current_time = datetime.now()
    expired_users = []
    
    for user_name, last_seen in list(active_sessions.items()):
        if (current_time - last_seen).total_seconds() > SESSION_EXPIRY_MINUTES * 60:
            expired_users.append(user_name)
    
    for user in expired_users:
        del active_sessions[user]
    
    # Build active users list
    active = []
    for user_name, last_seen in active_sessions.items():
        active.append({
            "user": user_name,
            "last_seen": str(last_seen),
            "status": "online"
        })
    
    return {
        "active_count": len(active),
        "users": active,
        "locked_records": {
            k: {
                "locked_by": v['user'],
                "locked_at": v['time']
            }
            for k, v in record_locks.items()
        }
    }

@app.post("/api/team/heartbeat")
async def heartbeat(request: Request):
    """Update user's active status"""
    user = request.headers.get("X-User", "anonymous")
    if user != "anonymous":
        active_sessions[user] = datetime.now()
    return {"status": "ok", "user": user}

# ============================================================================
# DEBUG ENDPOINTS
# ============================================================================

@app.get("/api/debug/sample-patients")
async def sample_patients(limit: int = 10):
    """Get sample patient hospital numbers for testing"""
    try:
        from app.database import emr_engine
        
        if not emr_engine:
            return {"error": "Database not configured"}
        
        with emr_engine.connect() as conn:
            rows = conn.execute(
                text("""
                    SELECT hospital_number, first_name, surname, uuid, sex
                    FROM patient_person 
                    WHERE archived = 0 AND hospital_number IS NOT NULL
                    LIMIT :lim
                """),
                {"lim": limit}
            ).fetchall()
            
            return {
                "patients": [
                    {
                        "hospital_number": r[0],
                        "name": f"{r[1]} {r[2]}",
                        "uuid": r[3],
                        "sex": r[4]
                    }
                    for r in rows
                ]
            }
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/debug/db-status")
async def database_status():
    """Check database connection status"""
    try:
        from app.database import emr_engine, dqa_engine
        
        status = {
            "emr": {"connected": False, "error": None},
            "dqa": {"connected": False, "error": None}
        }
        
        if emr_engine:
            try:
                with emr_engine.connect() as conn:
                    conn.execute(text("SELECT 1"))
                status["emr"]["connected"] = True
            except Exception as e:
                status["emr"]["error"] = str(e)
        
        if dqa_engine:
            try:
                with dqa_engine.connect() as conn:
                    conn.execute(text("SELECT 1"))
                status["dqa"]["connected"] = True
            except Exception as e:
                status["dqa"]["error"] = str(e)
        
        return status
        
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/debug/raw-regimens/{hospital_number:path}")
async def debug_raw_regimens(hospital_number: str):
    """Debug endpoint to see raw JSONB structure of refills"""
    try:
        from app.database import emr_engine
        
        hospital_number = unquote(hospital_number)
        
        with emr_engine.connect() as conn:
            result = conn.execute(
                text("SELECT uuid FROM patient_person WHERE hospital_number = :hn AND archived = 0"),
                {"hn": hospital_number}
            )
            row = result.fetchone()
            if not row:
                return {"error": "Patient not found"}
            
            person_uuid = row[0]
            
            refills = conn.execute(
                text("""
                    SELECT 
                        id,
                        visit_date::DATE,
                        extra,
                        extra -> 'regimens' AS regimens_raw,
                        jsonb_typeof(extra -> 'regimens') AS regimens_type,
                        jsonb_array_length(extra -> 'regimens') AS array_length
                    FROM hiv_art_pharmacy
                    WHERE person_uuid = :uuid AND archived = 0
                    AND extra -> 'regimens' IS NOT NULL
                    ORDER BY visit_date DESC
                    LIMIT 5
                """),
                {"uuid": person_uuid}
            ).fetchall()
            
            return {
                "patient_uuid": person_uuid,
                "refills": [
                    {
                        "id": r[0],
                        "visit_date": str(r[1]) if r[1] else None,
                        "extra_preview": str(r[2])[:500] if r[2] else None,
                        "regimens_type": r[4],
                        "array_length": r[5]
                    }
                    for r in refills
                ]
            }
    except Exception as e:
        return {"error": str(e)}

# ============================================================================
# ERROR HANDLERS
# ============================================================================

@app.exception_handler(404)
async def not_found_handler(request: Request, exc: HTTPException):
    """Handle 404 errors"""
    if request.url.path.startswith('/api/'):
        return JSONResponse(
            status_code=404,
            content={"success": False, "detail": "Endpoint not found"}
        )
    return HTMLResponse(
        content="""
        <html>
        <head><title>404 - Page Not Found</title>
        <style>
            body { font-family: 'Inter', sans-serif; display: flex; align-items: center; justify-content: center; height: 100vh; margin: 0; background: #f8fafc; }
            .error-card { text-align: center; padding: 3rem; background: white; border-radius: 16px; box-shadow: 0 10px 40px rgba(0,0,0,0.1); }
            h1 { font-size: 4rem; color: #4f46e5; margin: 0; }
            a { color: #4f46e5; text-decoration: none; font-weight: 600; }
        </style>
        </head>
        <body>
            <div class="error-card">
                <h1>404</h1>
                <p>Page not found</p>
                <a href='/'>Go to Home</a>
            </div>
        </body>
        </html>
        """,
        status_code=404
    )

@app.exception_handler(500)
async def internal_error_handler(request: Request, exc: HTTPException):
    """Handle 500 errors"""
    logger.error(f"500 error: {exc}")
    if request.url.path.startswith('/api/'):
        return JSONResponse(
            status_code=500,
            content={"success": False, "detail": "Internal server error"}
        )
    return HTMLResponse(
        content="""
        <html>
        <head><title>500 - Server Error</title>
        <style>
            body { font-family: 'Inter', sans-serif; display: flex; align-items: center; justify-content: center; height: 100vh; margin: 0; background: #f8fafc; }
            .error-card { text-align: center; padding: 3rem; background: white; border-radius: 16px; box-shadow: 0 10px 40px rgba(0,0,0,0.1); }
            h1 { font-size: 4rem; color: #ef4444; margin: 0; }
            a { color: #4f46e5; text-decoration: none; font-weight: 600; }
        </style>
        </head>
        <body>
            <div class="error-card">
                <h1>500</h1>
                <p>Internal server error</p>
                <a href='/'>Go to Home</a>
            </div>
        </body>
        </html>
        """,
        status_code=500
    )

# ============================================================================
# PHARMACY REPORT - PULLS FROM BOTH EMR (facility details) AND DQA (care card)
# ============================================================================

# ============================================================================
# PHARMACY EXCEL - SELF-CONTAINED (fetches its own data)
# ============================================================================

@app.get("/api/reports/pharmacy/excel")
async def generate_pharmacy_report_excel(
    request: Request,
    start_date: str = None,
    end_date: str = None
):
    """
    Generate Comprehensive DQA Match Report Excel
    """
    try:
        from app.database import dqa_engine, emr_engine
        from app.models.dqa_models import CareCardRecord
        from sqlalchemy.orm import Session
        from sqlalchemy import text
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        import io
        
        user = request.headers.get("X-User", "anonymous")
        logger.info(f"📊 User '{user}' generating comprehensive DQA report")
        
        if start_date and start_date.strip():
            start_date = start_date.strip()
        else:
            start_date = None
        if end_date and end_date.strip():
            end_date = end_date.strip()
        else:
            end_date = None
        
        facility_cache = {}
        
        def get_facility_info(hospital_number):
            if not hospital_number:
                return {"facility_name": "N/A", "datim_id": "N/A"}
            if hospital_number in facility_cache:
                return facility_cache[hospital_number]
            
            info = {"facility_name": "N/A", "datim_id": "N/A"}
            try:
                with emr_engine.connect() as conn:
                    result = conn.execute(text("""
                        SELECT facility.name, oi.code
                        FROM patient_person p
                        INNER JOIN hiv_enrollment h ON h.person_uuid = p.uuid AND h.archived = 0
                        INNER JOIN base_organisation_unit facility ON facility.id = h.facility_id
                        LEFT JOIN base_organisation_unit_identifier oi ON oi.organisation_unit_id = facility.id AND oi.name = 'DATIM_ID'
                        WHERE p.hospital_number = :hn AND p.archived = 0 LIMIT 1
                    """), {"hn": hospital_number}).fetchone()
                    if result:
                        info["facility_name"] = result[0] or 'N/A'
                        info["datim_id"] = result[1] or 'N/A'
            except Exception as e:
                logger.warning(f"EMR lookup failed for {hospital_number}: {e}")
            
            facility_cache[hospital_number] = info
            return info
        
        def get_status_display(results_list, field_name):
            if not results_list:
                return "—"
            for r in results_list:
                if r.get("field") == field_name:
                    if r.get("match"):
                        return "✅ Match"
                    corrected = r.get("corrected_on", "")
                    if corrected == "emr":
                        return "💻 Corrected (EMR)"
                    elif corrected == "care_card":
                        return "📋 Corrected (Card)"
                    elif corrected == "both":
                        return "🔀 Corrected (Both)"
                    return "⚠️ Mismatch"
            return "—"
        
        with Session(dqa_engine) as session:
            query = session.query(CareCardRecord).filter(CareCardRecord.is_verified == True)
            
            if start_date:
                query = query.filter(CareCardRecord.verified_at >= start_date)
            if end_date:
                query = query.filter(CareCardRecord.verified_at <= end_date + "T23:59:59")
            
            records = query.order_by(CareCardRecord.verified_at.desc()).all()
        
        logger.info(f"📊 Found {len(records)} verified records")
        
        report_data = []
        sno = 1
        
        for record in records:
            facility_info = get_facility_info(record.hospital_number)
            
            ed = record.enrollment_data or {}
            biodata_results = ed.get("biodata_verification", [])
            latest_refill_results = ed.get("latest_refill_verification", [])
            drug_pickups = record.drug_pickups or []
            viral_loads = record.viral_loads or []
            
            report_data.append({
                "s_no": sno,
                "facility_name": facility_info['facility_name'],
                "datim_id": facility_info['datim_id'],
                "patient_uuid": record.person_uuid or 'N/A',
                "hospital_number": record.hospital_number or 'N/A',
                "sex_status": get_status_display(biodata_results, "sex"),
                "dob_status": get_status_display(biodata_results, "date_of_birth"),
                "art_start_status": get_status_display(biodata_results, "art_start_date"),
                "pickup_status": get_status_display(latest_refill_results, "last_pickup_date"),
                "duration_status": get_status_display(drug_pickups, "refill_durations"),
                "vl_sample_date": get_status_display(viral_loads, "vl_sample_dates"),
                "vl_result_status": get_status_display(viral_loads, "vl_results"),
                "vl_result_date_status": get_status_display(viral_loads, "vl_result_dates"),
                "verified_by": record.verified_by or record.created_by or 'N/A',
                "verified_at": record.verified_at.date() if record.verified_at else None
            })
            sno += 1
        
        # ================================================================
        # CREATE EXCEL - Headers only, no title/summary rows
        # ================================================================
        wb = Workbook()
        ws = wb.active
        ws.title = "DQA Match Report"
        
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        match_fill = PatternFill(start_color="f0fdf4", end_color="f0fdf4", fill_type="solid")
        corrected_fill = PatternFill(start_color="fffbeb", end_color="fffbeb", fill_type="solid")
        mismatch_fill = PatternFill(start_color="fef2f2", end_color="fef2f2", fill_type="solid")
        not_reviewed_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='d1d5db'),
            right=Side(style='thin', color='d1d5db'),
            top=Side(style='thin', color='d1d5db'),
            bottom=Side(style='thin', color='d1d5db')
        )
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # ================================================================
        # HEADERS - Row 1
        # ================================================================
        headers = [
            "S/No.",
            "Facility",
            "DATIM ID",
            "Patient UUID",
            "Hospital Number",
            "Sex Status",
            "Date of Birth Status",
            "ART Start Status",
            "Pickup Status",
            "Month of ARV\n(Duration) Status",
            "Sample Collection\nDate Status",
            "VL Result\nStatus",
            "Date of Viral\nLoad Status",
            "Verified Date",
            "Verified by"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.row_dimensions[1].height = 36
        
        # ================================================================
        # DATA ROWS - Starting from Row 2
        # ================================================================
        for row_idx, row_data in enumerate(report_data):
            excel_row = 2 + row_idx
            
            values = [
                row_data['s_no'],
                row_data['facility_name'],
                row_data['datim_id'],
                row_data['patient_uuid'],
                row_data['hospital_number'],
                row_data['sex_status'],
                row_data['dob_status'],
                row_data['art_start_status'],
                row_data['pickup_status'],
                row_data['duration_status'],
                row_data['vl_sample_date'],
                row_data['vl_result_status'],
                row_data['vl_result_date_status'],
                row_data['verified_at'],
                row_data['verified_by']
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=excel_row, column=col, value=value)
                cell.border = thin_border
                cell.font = Font(size=9, name="Calibri")
                
                if col >= 6 and col <= 13:
                    cell.alignment = center_alignment
                    if "✅ Match" in str(value):
                        cell.fill = match_fill
                        cell.font = Font(size=9, color="059669", name="Calibri")
                    elif "Corrected" in str(value):
                        cell.fill = corrected_fill
                        cell.font = Font(size=9, color="d97706", name="Calibri")
                    elif "Mismatch" in str(value):
                        cell.fill = mismatch_fill
                        cell.font = Font(size=9, color="dc2626", name="Calibri")
                    elif str(value) == "—":
                        cell.fill = not_reviewed_fill
                        cell.font = Font(size=9, color="94a3b8", name="Calibri", italic=True)
                    elif col == 1:
                        cell.alignment = center_alignment
                    elif col == 14: 
                        cell.alignment = center_alignment
                        cell.font = Font(size=9, bold=True, name="Calibri", color="FF1e293b")
                    else:
                        cell.alignment = Alignment(vertical="center")
            
            ws.row_dimensions[excel_row].height = 22
        
        # ================================================================
        # COLUMN WIDTHS
        # ================================================================
        widths = {
            1: 6,    # S/No
            2: 28,   # Facility
            3: 14,   # DATIM ID
            4: 38,   # Patient UUID
            5: 18,   # Hospital Number
            6: 22,   # Sex Status
            7: 22,   # DOB Status
            8: 22,   # ART Start Status
            9: 22,   # Pickup Status
            10: 24,  # Duration Status
            11: 24,  # VL Sample Date
            12: 20,  # VL Result
            13: 22,   # VL Result Date
            14: 14,  # Verification Date
            15: 22   # Verified By
        }
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:O{1 + len(report_data)}"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"DQA_Match_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"📊 Report generation error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})

# ============================================================================
# VIRAL LOAD EXCEL - SELF-CONTAINED
# ============================================================================

@app.get("/api/reports/viral-load/excel")
async def generate_viral_load_report_excel(
    request: Request,
    start_date: str = None,
    end_date: str = None
):
    """Generate Viral Load Report Excel - fetches data directly"""
    try:
        from app.database import dqa_engine, emr_engine
        from app.models.dqa_models import CareCardRecord
        from sqlalchemy.orm import Session
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        user = request.headers.get("X-User", "anonymous")
        logger.info(f"User '{user}' generating VL Excel report")
        
        # Clean parameters
        if start_date and start_date.strip():
            start_date = start_date.strip()
        else:
            start_date = None
        
        if end_date and end_date.strip():
            end_date = end_date.strip()
        else:
            end_date = None
        
        # Cache for EMR data
        emr_cache = {}
        
        def get_emr_patient_info(hospital_number):
            if not hospital_number:
                return {"facility_name": "N/A", "datim_id": "N/A", "art_start_date": "N/A", "current_regimen": "N/A"}
            if hospital_number in emr_cache:
                return emr_cache[hospital_number]
            
            info = {"facility_name": "N/A", "datim_id": "N/A", "art_start_date": "N/A", "current_regimen": "N/A"}
            try:
                with emr_engine.connect() as conn:
                    result = conn.execute(text("""
                        SELECT facility.name, oi.code, h.date_of_registration, p.uuid
                        FROM patient_person p
                        INNER JOIN hiv_enrollment h ON h.person_uuid = p.uuid AND h.archived = 0
                        INNER JOIN base_organisation_unit facility ON facility.id = h.facility_id
                        LEFT JOIN base_organisation_unit_identifier oi ON oi.organisation_unit_id = facility.id AND oi.name = 'DATIM_ID'
                        WHERE p.hospital_number = :hn AND p.archived = 0 LIMIT 1
                    """), {"hn": hospital_number}).fetchone()
                    
                    if result:
                        info["facility_name"] = result[0] or 'N/A'
                        info["datim_id"] = result[1] or 'N/A'
                        info["art_start_date"] = str(result[2]) if result[2] else 'N/A'
                        
                        person_uuid = result[3]
                        if person_uuid:
                            cr = conn.execute(text("""
                                SELECT COALESCE(elem ->> 'regimenName', elem ->> 'name', 'Unknown')
                                FROM hiv_art_pharmacy hap
                                CROSS JOIN LATERAL jsonb_array_elements(hap.extra -> 'regimens') AS elem
                                WHERE hap.archived = 0 AND hap.person_uuid = :uuid
                                AND hap.extra -> 'regimens' IS NOT NULL
                                AND jsonb_typeof(hap.extra -> 'regimens') = 'array'
                                ORDER BY hap.visit_date DESC LIMIT 1
                            """), {"uuid": person_uuid}).fetchone()
                            if cr and cr[0]:
                                info["current_regimen"] = cr[0]
            except Exception as e:
                logger.warning(f"EMR lookup failed for {hospital_number}: {e}")
            
            emr_cache[hospital_number] = info
            return info
        
        # Fetch DQA data
        with Session(dqa_engine) as session:
            query = session.query(CareCardRecord)
            if start_date:
                query = query.filter(CareCardRecord.verified_at >= start_date)
            if end_date:
                query = query.filter(CareCardRecord.verified_at <= end_date)
            records = query.order_by(CareCardRecord.updated_at.desc()).all()
        
        # Build VL data
        vl_data = []
        sno = 1
        
        for record in records:
            emr_info = get_emr_patient_info(record.hospital_number)
            
            for vl in (record.viral_loads or []):
                vl_result = vl.get('viral_load_result', 'N/A') or 'N/A'
                
                vl_data.append({
                    "s_no": sno,
                    "facility_name": emr_info['facility_name'],
                    "datim_id": emr_info['datim_id'],
                    "patient_id": record.person_uuid or 'N/A',
                    "hospital_num": record.hospital_number or 'N/A',
                    "sample_collection_date": vl.get('sample_collection_date', 'N/A') or 'N/A',
                    "viral_load_result": vl_result,
                    "result_date": vl.get('result_date', 'N/A') or 'N/A',
                    "vl_classification": classify_vl_result(vl_result),
                    "art_start_date": emr_info['art_start_date'],
                    "current_regimen": emr_info['current_regimen']
                })
                sno += 1
        
        logger.info(f"VL Excel: {len(vl_data)} rows")
        
        # Create Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Viral Load Report"
        
        headers = [
            "S/No", "Facility Name", "DATIM Id", "Patient Id", "Hospital Num",
            "Sample Collection Date", "Viral Load Result", "Result Date",
            "VL Classification", "ART Start Date", "Current Regimen"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        for row_idx, row_data in enumerate(vl_data, 2):
            ws.cell(row=row_idx, column=1, value=row_data['s_no']).border = thin_border
            ws.cell(row=row_idx, column=2, value=row_data['facility_name']).border = thin_border
            ws.cell(row=row_idx, column=3, value=row_data['datim_id']).border = thin_border
            ws.cell(row=row_idx, column=4, value=row_data['patient_id']).border = thin_border
            ws.cell(row=row_idx, column=5, value=row_data['hospital_num']).border = thin_border
            ws.cell(row=row_idx, column=6, value=row_data['sample_collection_date']).border = thin_border
            ws.cell(row=row_idx, column=7, value=row_data['viral_load_result']).border = thin_border
            ws.cell(row=row_idx, column=8, value=row_data['result_date']).border = thin_border
            ws.cell(row=row_idx, column=9, value=row_data['vl_classification']).border = thin_border
            ws.cell(row=row_idx, column=10, value=row_data['art_start_date']).border = thin_border
            ws.cell(row=row_idx, column=11, value=row_data['current_regimen']).border = thin_border
        
        widths = {1:6, 2:30, 3:15, 4:38, 5:18, 6:20, 7:15, 8:20, 9:18, 10:20, 11:35}
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        
        ws.freeze_panes = 'A2'
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=VL_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"VL Excel error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})

# ============================================================================
# HELPER FUNCTIONS FOR REPORTS
# ============================================================================

def categorize_regimen(name):
    """Categorize a regimen based on its drug components"""
    if not name:
        return 'ARVs'
    
    name_lower = str(name).lower()
    
    # Anti-TB drugs
    tb_drugs = ['isoniazid', 'inh', '3hp', 'rifampicin', 'rifampin', 'rifapentine',
                'pyrazinamide', 'ethambutol', 'rhze', 'rhz', 'tb ', 'rifinah', 'akurit']
    for tb in tb_drugs:
        if tb in name_lower:
            return 'Anti-TB'
    
    # Prophylaxis drugs
    prophylaxis_drugs = ['cotrimoxazole', 'septrin', 'bactrim', 'fluconazole',
                         'dapsone', 'azithromycin', 'nystatin', 'clotrimazole']
    for proph in prophylaxis_drugs:
        if proph in name_lower:
            return 'Prophylaxis'
    
    # Other drugs
    other_drugs = ['pyridoxine', 'vitamin', 'folic', 'ferrous', 'paracetamol',
                   'ibuprofen', 'amitriptyline', 'amoxicillin']
    for other in other_drugs:
        if other in name_lower:
            return 'Other'
    
    # ARV components
    arv_components = ['tdf', '3tc', 'dtg', 'efv', 'nvp', 'abc', 'azt', 'lpv',
                      'atv', 'taf', 'ftc', 'd4t', 'tenofovir', 'lamivudine',
                      'dolutegravir', 'efavirenz', 'nevirapine', 'abacavir',
                      'zidovudine', 'lopinavir', 'atazanavir']
    for arv in arv_components:
        if arv in name_lower:
            return 'ARVs'
    
    return 'ARVs'


def get_mmd_type(duration):
    """Get MMD type based on duration"""
    if not duration:
        return 'N/A'
    duration = int(duration)
    if duration <= 30:
        return 'MMD-1'
    elif duration <= 60:
        return 'MMD-2'
    elif duration <= 90:
        return 'MMD-3'
    elif duration <= 120:
        return 'MMD-4'
    else:
        return 'MMD-6'


def classify_vl_result(vl_result):
    """Classify viral load result"""
    if not vl_result:
        return 'Unknown'
    
    vl_str = str(vl_result).strip()
    
    if vl_str.startswith('<'):
        return 'Suppressed'
    
    try:
        vl_num = int(float(vl_str.replace(',', '')))
        if vl_num < 200:
            return 'Suppressed'
        elif vl_num < 1000:
            return 'Low Viremia'
        else:
            return 'Unsuppressed'
    except (ValueError, TypeError):
        return 'Unknown'


def get_current_regimen_from_record(record):
    """Get the most recent regimen from a care card record"""
    drug_pickups = record.drug_pickups or []
    if not drug_pickups:
        return 'N/A'
    
    # Get the first (most recent) pickup
    first_pickup = drug_pickups[0]
    return first_pickup.get('regimen', '') or first_pickup.get('regimen_name', '') or 'N/A'



# ============================================================================
# LAB SETTINGS API
# ============================================================================

@app.get("/api/lab/settings")
async def get_lab_settings(request: Request):
    """Get current lab settings"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import LabSettings
        from sqlalchemy.orm import Session
        
        with Session(dqa_engine) as session:
            settings = session.query(LabSettings).order_by(LabSettings.id.desc()).first()
            
            if not settings:
                settings = LabSettings()
                session.add(settings)
                session.commit()
            
            return {
                "success": True,
                "data": {
                    "id": settings.id,
                    "pcr_lab_name": settings.pcr_lab_name or "",
                    "facility_name": settings.facility_name or "",
                    "clinician_name": settings.clinician_name or "",
                    "assayed_by_name": settings.assayed_by_name or "",
                    "approved_by_name": settings.approved_by_name or "",
                    "collected_by_name": settings.collected_by_name or "",
                    "auto_fill_dates": settings.auto_fill_dates if hasattr(settings, 'auto_fill_dates') else True
                }
            }
    except Exception as e:
        logger.error(f"Lab settings error: {e}")
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})


@app.put("/api/lab/settings")
async def update_lab_settings(request: Request):
    """Update lab settings"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import LabSettings
        from sqlalchemy.orm import Session
        
        data = await request.json()
        user = request.headers.get("X-User", "anonymous")
        
        with Session(dqa_engine) as session:
            settings = session.query(LabSettings).order_by(LabSettings.id.desc()).first()
            
            if not settings:
                settings = LabSettings()
                session.add(settings)
            
            if 'pcr_lab_name' in data:
                settings.pcr_lab_name = data['pcr_lab_name']
            if 'facility_name' in data:
                settings.facility_name = data['facility_name']
            if 'clinician_name' in data:
                settings.clinician_name = data['clinician_name']
            if 'assayed_by_name' in data:
                settings.assayed_by_name = data['assayed_by_name']
            if 'approved_by_name' in data:
                settings.approved_by_name = data['approved_by_name']
            if 'collected_by_name' in data:
                settings.collected_by_name = data['collected_by_name']
            
            settings.updated_by = user
            settings.updated_at = datetime.now()
            
            session.commit()
            
            return {"success": True, "message": "Lab settings updated successfully"}
            
    except Exception as e:
        logger.error(f"Update lab settings error: {e}")
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})


# ============================================================================
# VL RESULT FORMATTING & PDF GENERATION - COMPLETE FULL CODE
# ALL BOXES PROPERLY PLACED - 100% COMPLETE
# ============================================================================

def clean_regimen_name(name):
    """
    Clean regimen name for display.
    Converts "TDF(300mg)/3TC(300mg)/DTG(50mg)" → "TDF-3TC-DTG"
    Removes dosage in parentheses and replaces / with -
    """
    if not name or name == "Unknown":
        return "N/A"
    
    import re
    cleaned = re.sub(r'\([^)]*\)', '', name)
    cleaned = cleaned.replace('/', '-')
    cleaned = re.sub(r'-+', '-', cleaned)
    cleaned = cleaned.strip('-').strip()
    
    return cleaned if cleaned else "N/A"


def format_vl_for_print(vl_value):
    """
    Format VL result for printing.
    - 0 or TND → "TND"
    - 1-20 → "<20"
    - 21+ → actual number
    """
    if not vl_value or vl_value == 'N/A':
        return 'N/A'
    
    vl_str = str(vl_value).strip().upper()
    
    if vl_str in ['TND', 'NOT DETECTED', 'UNDETECTABLE', '0', '0.0']:
        return 'TND'
    
    if vl_str.startswith('<'):
        try:
            num = int(float(vl_str[1:].strip().replace(',', '')))
            if num == 0: return 'TND'
            if num <= 20: return '<20'
            return str(num)
        except: return 'N/A'
    
    try:
        num = int(float(vl_str.replace(',', '')))
        if num == 0: return 'TND'
        if num <= 20: return '<20'
        return str(num)
    except: return 'N/A'


def create_vl_pdf(data):
    """
    Create VL Result PDF matching UDUTH form template.
    100% Complete - All boxes properly placed.
    """
    import tempfile
    import os
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    
    path = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False).name
    c = canvas.Canvas(path, pagesize=A4)
    width, height = A4

    def box(x, y, w, h):
        """Draw a rectangle box"""
        c.rect(x, y, w, h, stroke=1, fill=0)

    def text(txt, x, y, size=9, bold=False):
        """Draw text at position"""
        if txt is None: txt = ""
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawString(x, y, str(txt))

    def text_center(txt, y, size=9, bold=False):
        """Draw centered text"""
        if txt is None: txt = ""
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawCentredString(width / 2, y, str(txt))

    def text_right(txt, x, y, size=9, bold=False):
        """Draw right-aligned text"""
        if txt is None: txt = ""
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawRightString(x, y, str(txt))

    # ========================================================================
    # START POSITION
    # ========================================================================
    y = height - 35

    # ========================================================================
    # HEADER WITH LOGO
    # ========================================================================
    logo_path = resource_path(os.path.join("app", "static", "logo.png"))
    
    if os.path.exists(logo_path):
        try:
            logo = ImageReader(logo_path)
            c.drawImage(logo, 35, y - 30, width=50, height=42, preserveAspectRatio=True, mask='auto')
        except:
            box(35, y - 30, 50, 42)
    else:
        box(35, y - 30, 50, 42)

    # Title centered on page
    text_center("VIRAL LOAD ORDER AND RESULTS FORM", y, 16, True)
    text_center(data.get("pcr_lab_name", "UDUTH Testing Lab"), y - 18, 12, True)
    
    y -= 55

    # ========================================================================
    # FACILITY NAME & STATE
    # ========================================================================
    text("FACILITY NAME:", 40, y, 9, True)
    box(130, y - 12, 320, 16)
    text(data.get("facility_name", ""), 135, y - 10, 8)
    
    y -= 25

    text("STATE:", 40, y, 9, True)
    box(130, y - 12, 320, 16)
    text(data.get("state", ""), 135, y - 10, 8)
    
    y -= 35

    # ========================================================================
    # PATIENT NAME - SURNAME + OTHER NAMES
    # ========================================================================
    text("PATIENT NAME", 40, y, 9, True)

    # Surname box
    box(40, y - 22, 250, 18)
    text(data.get("surname", ""), 45, y - 19, 9)
    text("Surname", 45, y - 35, 7)

    # Other names box
    other_names = data.get("other_names", "")
    box(300, y - 22, 250, 18)
    text(other_names, 305, y - 19, 9)
    text("Other name(s)", 305, y - 35, 7)
    
    y -= 48

    # ========================================================================
    # SEX / AGE / PREGNANT / BREASTFEEDING - ALL IN ONE ROW
    # ========================================================================
    row_y = y

    # Male checkbox
    box(40, row_y - 4, 12, 12)
    if str(data.get("gender", "")).lower() == "male":
        text("X", 43, row_y - 6, 10, True)
    text("Male", 56, row_y - 3, 8)

    # Female checkbox
    box(100, row_y - 4, 12, 12)
    if str(data.get("gender", "")).lower() == "female":
        text("X", 103, row_y - 6, 10, True)
    text("Female", 116, row_y - 3, 8)

    # Age
    text("Age", 170, row_y - 3, 8)
    box(195, row_y - 6, 35, 16)
    text(data.get("age", ""), 200, row_y - 3, 9)
    text("Years", 235, row_y - 3, 8)

    # Under 2 years
    box(280, row_y - 4, 12, 12)
    text("< 2 years", 296, row_y - 3, 7)

    # Months
    box(350, row_y - 6, 30, 16)
    text("Months", 384, row_y - 3, 7)

    # Pregnant
    box(430, row_y - 4, 12, 12)
    text("Pregnant", 446, row_y - 3, 7)

    # Breastfeeding
    box(500, row_y - 4, 12, 12)
    text("Breastfeeding", 516, row_y - 3, 7)

    y -= 35

    # ========================================================================
    # IDENTIFICATION NUMBERS
    # ========================================================================
    text("ID", 40, y, 9, True)
    box(40, y - 20, 180, 18)
    text(data.get("unique_id", data.get("hospital_number", "")), 45, y - 17, 9)
    text("Client's Unique Number", 45, y - 34, 7)

    text("HOSPITAL (UNIT) NO.", 235, y, 9, True)
    box(235, y - 20, 180, 18)
    text(data.get("hospital_number", ""), 240, y - 17, 9)

    text("LAB REGISTRATION NO.", 430, y, 9, True)
    box(430, y - 20, 120, 18)
    text(str(data.get("lab_sample_no", "")), 435, y - 17, 9)
    
    y -= 45

    # ========================================================================
    # SAMPLE TYPE
    # ========================================================================
    text("SAMPLE TYPE:", 40, y, 9, True)
    text("Plasma", 140, y, 9)
    
    y -= 25

    # ========================================================================
    # INDICATION FOR VL TEST BOX
    # ========================================================================
    box(40, y - 105, width - 80, 105)
    
    text("INDICATION FOR VIRAL LOAD TEST", 50, y - 20, 9, True)
    text("Routine (every >12 months)", 50, y - 40, 8)
    text("Others", 50, y - 55, 8)

    # ART Commencement Date
    text("ART Commencement Date", 320, y - 20, 9, True)
    box(320, y - 40, 150, 18)
    text(data.get("art_start_date", ""), 325, y - 36, 8)
    text("DD/MM/YY", 360, y - 55, 7)

    # Drug Regimen
    text("DRUG REGIMEN", 320, y - 65, 9, True)
    regimen = data.get("regimen", "N/A")
    text(regimen, 320, y - 82, 9)

    y -= 120

    # ========================================================================
    # DISCLAIMER
    # ========================================================================
    c.setFont("Helvetica", 7)
    c.drawString(40, y, "Tests will only be performed at the oratory, if all fields above are filled in correctly and signed below by the ordering staff")
    
    y -= 18

    # ========================================================================
    # COLLECTED BY & SIGNATURE
    # ========================================================================
    text("Collected by:", 40, y, 9, True)
    text(data.get("collected_by", ""), 125, y, 9)
    
    c.setFont("Helvetica", 7)
    c.drawCentredString(220, y - 28, "Print Name")

    text("Signature", 370, y, 9, True)
    
    y -= 35

    # Collection Date & Time
    text("Collection Date:", 370, y, 9, True)
    text(data.get("collection_date", ""), 470, y, 9)

    text("Collection Time:", 370, y - 15, 9, True)
    text(data.get("collection_time", "10:00 AM"), 470, y - 15, 9)

    y -= 45

    # ========================================================================
    # HORIZONTAL SEPARATOR
    # ========================================================================
    c.setLineWidth(1.5)
    c.line(40, y, width - 40, y)
    
    y -= 25

    # ========================================================================
    # RESULTS SECTION
    # ========================================================================
    text_center("RESULTS", y, 13, True)
    
    y -= 30

    # Date Received & Time
    text("DATE RECEIVED AT PCR LAB", 40, y, 9, True)
    text(data.get("received_date", ""), 190, y, 9)

    text("TIME", 370, y, 9, True)
    text(data.get("received_time", "10:00 AM"), 470, y, 9)

    y -= 30

    # RESULT VALUE + PCR LAB NAME (same row)
    text("RESULTS", 40, y, 9, True)
    box(90, y - 16, 180, 22)
    
    result_value = data.get("result_value", "N/A")
    if result_value in ['TND', 'N/A']:
        display_text = result_value
    else:
        display_text = f"{result_value} copies/ml"
    
    text(display_text, 100, y - 11, 11, True)

    text("PCR LAB NAME", 320, y, 9, True)
    text(data.get("pcr_lab_name", "UDUTH Testing Lab"), 420, y, 9)

    y -= 35

    # PCR LAB SAMPLE NO
    text("PCR LAB SAMPLE NO.", 320, y, 9, True)
    text(str(data.get("", "")), 450, y, 9)

    y -= 25

    # ========================================================================
    # RESULT INTERPRETATION BOXES
    # ========================================================================
    box(40, y - 40, 130, 35)
    text("Result Interpretation", 50, y - 28, 8, True)

    box(210, y - 40, 120, 35)
    c.setFont("Helvetica", 7)
    c.drawString(215, y - 28, "Viral suppression")
    c.drawString(215, y - 38, "<1000 c/ml")

    box(350, y - 40, 120, 35)
    c.setFont("Helvetica", 7)
    c.drawString(355, y - 28, "Poor suppression")
    c.drawString(355, y - 38, "1000-10000 c/ml")

    box(490, y - 40, 100, 35)
    c.setFont("Helvetica", 7)
    c.drawString(495, y - 28, "Critical values")
    c.drawString(495, y - 38, ">10000 c/ml")

    y -= 60

        # ========================================================================
    # SIGNATURES SECTION - COMPACT WITH NAME + DATE ON SAME LINE
    # ========================================================================
    box(40, y - 55, width - 80, 55)
    
    # Vertical dividers
    c.line(210, y, 210, y - 55)
    c.line(380, y, 380, y - 55)

    # Names and dates on ONE line
    clinician_date = data.get("clinician_date", data.get("collection_date", ""))
    assayed_date = data.get("assayed_date", data.get("result_date", ""))
    approved_date = data.get("approved_date", data.get("result_date", ""))

    c.setFont("Helvetica-Bold", 7)
    c.drawString(50, y - 15, f"{data.get('clinician_name', '')}")
    c.drawString(220, y - 15, f"{data.get('assayed_by_name', '')}")
    c.drawString(390, y - 15, f"{data.get('approved_by_name', '')}")
    y -=15

    c.setFont("Helvetica", 7)
    c.drawString(100, y, f"{clinician_date}")
    c.drawString(280, y, f"{assayed_date}")
    c.drawString(490, y, f"{approved_date}")
    y += 20
    # Column headers at the bottom
    c.setFont("Helvetica", 9)
    c.drawString(55, y - 38, "Clinician-Date")
    c.drawString(235, y - 38, "Assayed By-Date")
    c.drawString(395, y - 38, "Reviewed-Approved-Date")

    # ========================================================================
    # FOOTER
    # ========================================================================
    c.setFont("Helvetica", 7)
    c.drawCentredString(width / 2, 30, "Copyright © 2020 Federal Ministry of Health, Nigeria. All rights reserved.")

    c.save()
    return path


# ============================================================================
# VL RESULT PRINT API
# ============================================================================

@app.get("/api/vl/print/{hospital_number:path}/{sample_date}")
async def print_vl_result(hospital_number: str, sample_date: str, request: Request):
    """Generate VL Result PDF"""
    try:
        from app.database import emr_engine, dqa_engine
        from app.models.dqa_models import LabSettings
        from sqlalchemy.orm import Session
        
        hospital_number = unquote(hospital_number)
        sample_date = unquote(sample_date)
        
        logger.info(f"🖨️ Printing VL result for {hospital_number} on {sample_date}")
        
        patient_result = None
        vl_result = None
        current_regimen = "N/A"
        
        # ALL EMR QUERIES INSIDE ONE WITH BLOCK
        with emr_engine.connect() as conn:
            # Get patient info
            patient_result = conn.execute(
                text("""
                    SELECT 
                        p.first_name, p.surname, p.other_name,
                        INITCAP(p.sex) AS sex,
                        EXTRACT(YEAR FROM AGE(CURRENT_DATE, p.date_of_birth)) AS age,
                        p.date_of_birth, p.hospital_number,
                        hac.visit_date::DATE AS art_start_date,
                        facility.name AS facility_name,
                        facility_state.name AS state,
                        h.unique_id
                    FROM patient_person p
                    INNER JOIN hiv_enrollment h ON h.person_uuid = p.uuid AND h.archived = 0
                    INNER JOIN base_organisation_unit facility ON facility.id = h.facility_id
                    INNER JOIN base_organisation_unit facility_lga ON facility_lga.id = facility.parent_organisation_unit_id
                    INNER JOIN base_organisation_unit facility_state ON facility_state.id = facility_lga.parent_organisation_unit_id
                    LEFT JOIN hiv_art_clinical hac ON hac.hiv_enrollment_uuid = h.uuid 
                        AND hac.archived = 0 AND hac.is_commencement = TRUE
                     WHERE p.hospital_number = :hn AND p.archived = 0 LIMIT 1
                """),
                {"hn": hospital_number}
            ).fetchone()
            
            # Get VL result
            vl_result = conn.execute(
                text("""
                    SELECT 
                        CAST(ls.date_sample_collected AS DATE) AS sample_date,
                        sm.result_reported AS vl_result,
                        CAST(sm.date_result_reported AS DATE) AS result_date,
                        COALESCE(ls.sample_number, ls.uuid::text) AS lab_sample_no
                    FROM laboratory_result sm
                    INNER JOIN laboratory_test lt ON lt.id = sm.test_id
                    INNER JOIN laboratory_sample ls ON ls.test_id = lt.id
                    WHERE lt.lab_test_id = 16
                    AND sm.patient_uuid = (SELECT uuid FROM patient_person WHERE hospital_number = :hn AND archived = 0)
                    AND CAST(ls.date_sample_collected AS DATE) = CAST(:sd AS DATE)
                    AND sm.result_reported IS NOT NULL AND sm.archived = 0
                    LIMIT 1
                """),
                {"hn": hospital_number, "sd": sample_date}
            ).fetchone()
            
            # Get current regimen
            try:
                cr = conn.execute(
                    text("""
                        SELECT elem.value ->> 'regimenName' AS regimen
                        FROM hiv_art_pharmacy hap
                        CROSS JOIN LATERAL jsonb_array_elements(hap.extra -> 'regimens') WITH ORDINALITY AS elem(value, ordinality)
                        WHERE hap.person_uuid = (SELECT uuid FROM patient_person WHERE hospital_number = :hn AND archived = 0)
                        AND hap.archived = 0 AND elem.ordinality = 1
                        ORDER BY hap.visit_date DESC LIMIT 1
                    """),
                    {"hn": hospital_number}
                ).fetchone()
                if cr and cr[0]:
                    current_regimen = cr[0]
                logger.info(f"✅ Current regimen for PDF: {current_regimen}")
            except Exception as e:
                logger.warning(f"Could not get regimen for PDF: {e}")
        
        if not patient_result:
            return JSONResponse(status_code=404, content={"success": False, "detail": "Patient not found"})
        
        if not vl_result:
            return JSONResponse(status_code=404, content={"success": False, "detail": f"VL result not found for date: {sample_date}"})
        
        # Get lab settings
        lab_settings = None
        try:
            with Session(dqa_engine) as session:
                lab_settings = session.query(LabSettings).order_by(LabSettings.id.desc()).first()
        except:
            pass
        
        # Format VL result
        vl_value = str(vl_result[1]).strip() if vl_result[1] else 'N/A'
        vl_display = format_vl_for_print(vl_value)
        
        # Generate lab numbers
        lab_reg = abs(hash(str(patient_result[6]) + str(sample_date))) % 100000
        lab_sample_no = str(vl_result[3]) if vl_result[3] else str(abs(hash(str(vl_result[0]) + str(vl_result[1]))) % 100000)
        
        
        # Build PDF data
        pdf_data = {
            "facility_name": (lab_settings.facility_name if lab_settings and lab_settings.facility_name else (patient_result[8] or 'N/A')),
            "state": patient_result[9] or 'N/A',
            "surname": patient_result[1] or '',
            "first_name": patient_result[0] or '',
            "other_name": patient_result[2] or '',
            "other_names": f"{patient_result[0] or ''} {patient_result[2] or ''}".strip(),
            "gender": patient_result[3] or '',
            "age": str(int(patient_result[4])) if patient_result[4] else '',
            "hospital_number": patient_result[6] or '',
            "unique_id": patient_result[10] or '',
            "art_start_date": str(patient_result[7])[:10] if patient_result[7] else '',
            "regimen": clean_regimen_name(current_regimen),
            "collection_date": str(vl_result[0]) if vl_result[0] else '',
            "result_value": vl_display,
            "received_date": str(vl_result[0]) if vl_result[0] else '',
            "received_time": "10:00 AM",
            "result_date": str(vl_result[2]) if vl_result[2] else '',
            "lab_reg_no": str(lab_reg),
            "lab_sample_no": lab_sample_no,
            "pcr_lab_name": lab_settings.pcr_lab_name if lab_settings else "UDUTH Testing Lab",
            "clinician_name": lab_settings.clinician_name if lab_settings else "",
            "assayed_by_name": lab_settings.assayed_by_name if lab_settings else "",
            "approved_by_name": lab_settings.approved_by_name if lab_settings else "",
            "clinician_date": str(vl_result[0]) if vl_result[0] else '',
            "assayed_date": str(vl_result[2]) if vl_result[2] else '',
            "approved_date": str(vl_result[2]) if vl_result[2] else '',
            "collected_by": lab_settings.collected_by_name if lab_settings else "",
            "collection_time": "10:00 AM",
        }
        
        logger.info(f"PDF data regimen: {pdf_data['regimen']}")
        
        # Generate PDF
        pdf_path = create_vl_pdf(pdf_data)
        
        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=f"VL_Result_{hospital_number.replace('/', '_')}_{sample_date}.pdf"
        )
        
    except Exception as e:
        logger.error(f"Print VL error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})
    


# ============================================================================
# USER AUTHENTICATION - SESSION STORE (defined early; used by middleware)
# ============================================================================

user_sessions: dict = {}  # token -> {user_id, username, full_name, role, expires}


def hash_password(password: str) -> str:
    """Simple password hashing"""
    salt = "meddqa_salt_2024"
    return hashlib.sha256(f"{password}{salt}".encode()).hexdigest()


def verify_password(password: str, password_hash: str) -> bool:
    """Verify password against hash"""
    return hash_password(password) == password_hash


def get_current_user_from_token(token: str):
    """Get current user from session token; returns None if missing/expired."""
    if token and token in user_sessions:
        session = user_sessions[token]
        if session["expires"] > datetime.now():
            return session
    return None


def log_activity(session, user_id, username, action, details, hospital_number, ip_address):
    """Log user activity"""
    try:
        from app.models.dqa_models import ActivityLog
        log = ActivityLog(
            user_id=user_id,
            username=username,
            action=action,
            details=details or {},
            hospital_number=hospital_number,
            ip_address=ip_address
        )
        session.add(log)
    except Exception as e:
        logger.warning(f"Activity log error: {e}")


@app.middleware("http")
async def combined_middleware(request: Request, call_next):
    """Combined middleware: setup check + authentication"""
    
    # ========================================================================
    # 1. PATHS THAT DON'T REQUIRE SETUP CHECK
    # ========================================================================
    setup_free_paths = [
        '/setup', '/api/setup', '/static', '/health',
        '/api/debug', '/api/schema', '/favicon.ico',
        '/login', '/api/auth'  # ✅ All auth paths
    ]
    
    needs_setup_check = not any(request.url.path.startswith(p) for p in setup_free_paths)
    
    if needs_setup_check and not is_setup_complete():
        return RedirectResponse(url='/setup', status_code=302)
    
    # ========================================================================
    # 2. PATHS THAT DON'T REQUIRE AUTHENTICATION
    # ========================================================================
    public_paths = [
        '/login', 
        '/api/auth',      # ✅ This covers /api/auth/login, /api/auth/logout, /api/auth/me
        '/setup', 
        '/api/setup', 
        '/static', 
        '/health',
        '/api/debug', 
        '/favicon.ico'
    ]
    
    is_public = any(request.url.path.startswith(p) for p in public_paths)
    
    if is_public:
        # Track user for backward compatibility
        user = request.headers.get("X-User", "anonymous")
        if user and user != "anonymous":
            active_sessions[user] = datetime.now()
        return await call_next(request)
    
    # ========================================================================
    # 3. AUTHENTICATION CHECK FOR PROTECTED PATHS
    # ========================================================================
    token = request.headers.get("X-Session-Token") or request.cookies.get("meddqa_token")
    
    if token and token in user_sessions:
        session = user_sessions[token]
        if session["expires"] > datetime.now():
            # Extend session
            session["expires"] = datetime.now() + timedelta(hours=12)
            return await call_next(request)
    
    # ========================================================================
    # 4. NOT AUTHENTICATED
    # ========================================================================
    if not request.url.path.startswith('/api/'):
        return RedirectResponse(url='/login', status_code=302)
    
    return JSONResponse(
        status_code=401,
        content={"success": False, "detail": "Authentication required"}
    )
    # ============================================================================
# USER AUTHENTICATION API
# ============================================================================


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """Login page"""
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/api/auth/login")
async def login(request: Request):
    """User login"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import User
        from sqlalchemy.orm import Session
        
        data = await request.json()
        username = data.get('username', '').strip()
        password = data.get('password', '')
        ip_address = request.client.host if request.client else 'unknown'
        
        logger.info(f"Login attempt: {username} from {ip_address}")
        
        if not username or not password:
            return JSONResponse(status_code=400, content={
                "success": False, "detail": "Username and password required"
            })
        
        with Session(dqa_engine) as session:
            user = session.query(User).filter(
                User.username == username,
                User.is_active == True
            ).first()
            
            if not user:
                logger.warning(f"Login failed: user '{username}' not found")
                return JSONResponse(status_code=401, content={
                    "success": False, "detail": "Invalid username or password"
                })
            
            if not verify_password(password, user.password_hash):
                logger.warning(f"Login failed: invalid password for '{username}'")
                log_activity(session, user.id, username, "login_failed", 
                            {"reason": "invalid_password"}, None, ip_address)
                return JSONResponse(status_code=401, content={
                    "success": False, "detail": "Invalid username or password"
                })
            
            # Generate session token
            session_token = secrets.token_urlsafe(32)
            user_sessions[session_token] = {
                "user_id": user.id,
                "username": user.username,
                "full_name": user.full_name,
                "role": user.role,
                "expires": datetime.now() + timedelta(hours=12)
            }
            
            # Update last login
            user.last_login = datetime.now()
            
            # Log activity
            log_activity(session, user.id, username, "login", 
                        {"ip": ip_address}, None, ip_address)
            
            session.commit()
            
            logger.info(f"✅ Login successful: {username} ({user.role})")
            
            return {
                "success": True,
                "token": session_token,
                "user": {
                    "id": user.id,
                    "full_name": user.full_name,
                    "username": user.username,
                    "role": user.role,
                    "position": user.position or ''
                }
            }
            
    except Exception as e:
        logger.error(f"Login error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})


@app.post("/api/auth/logout")
async def logout(request: Request):
    """User logout"""
    token = request.headers.get("X-Session-Token", "")
    if token in user_sessions:
        user = user_sessions[token]
        logger.info(f"Logout: {user['username']}")
        del user_sessions[token]
    return {"success": True, "message": "Logged out successfully"}


@app.get("/api/auth/me")
async def get_current_user(request: Request):
    """Get current logged-in user"""
    token = request.headers.get("X-Session-Token", "")
    
    session = get_current_user_from_token(token)
    
    if not session:
        return JSONResponse(status_code=401, content={
            "success": False, "detail": "Not authenticated"
        })
    
    return {
        "success": True,
        "user": {
            "id": session["user_id"],
            "full_name": session["full_name"],
            "username": session["username"],
            "role": session["role"]
        }
    }

@app.on_event("startup")
async def startup_event():
    """Initialize application on startup"""
    logger.info("=" * 60)
    logger.info("🚀 MedDQA v3.0 Starting Up...")
    logger.info(f"Setup complete: {is_setup_complete()}")
    
    if is_setup_complete():
        try:
            from app.database import emr_engine, dqa_engine
            if emr_engine:
                logger.info("✅ EMR Database connected")
            if dqa_engine:
                logger.info("✅ DQA Database connected")
                
                # ✅ Create default admin if no users exist
                try:
                    from app.models.dqa_models import User
                    from sqlalchemy.orm import Session
                    
                    with Session(dqa_engine) as session:
                        user_count = session.query(User).count()
                        if user_count == 0:
                            admin = User(
                                full_name="System Administrator",
                                username="admin",
                                password_hash=hash_password("password"),
                                role="admin",
                                position="Administrator"
                            )
                            session.add(admin)
                            session.commit()
                            logger.info("✅ Default admin created: admin/password")
                except Exception as e:
                    logger.warning(f"Default admin check: {e}")
        except Exception as e:
            logger.warning(f"Database initialization: {e}")
    
    logger.info("=" * 60)
# ============================================================================
# USER MANAGEMENT API (Admin Only)
# ============================================================================

@app.get("/api/users")
async def list_users(request: Request):
    """List all users (Admin only)"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import User
        from sqlalchemy.orm import Session
        
        token = request.headers.get("X-Session-Token", "")
        session_data = get_current_user_from_token(token)
        
        if not session_data or session_data.get("role") != "admin":
            return JSONResponse(status_code=403, content={
                "success": False, "detail": "Admin access required"
            })
        
        with Session(dqa_engine) as session:
            users = session.query(User).order_by(User.created_at.desc()).all()
            
            return {
                "success": True,
                "users": [
                    {
                        "id": u.id,
                        "full_name": u.full_name,
                        "username": u.username,
                        "role": u.role,
                        "position": u.position or '',
                        "is_active": u.is_active,
                        "last_login": str(u.last_login)[:19] if u.last_login else None,
                        "created_at": str(u.created_at)[:19] if u.created_at else None
                    }
                    for u in users
                ]
            }
    except Exception as e:
        logger.error(f"List users error: {e}")
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})


@app.post("/api/users")
async def create_user(request: Request):
    """Create new user (Admin only)"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import User
        from sqlalchemy.orm import Session
        
        token = request.headers.get("X-Session-Token", "")
        session_data = get_current_user_from_token(token)
        
        if not session_data or session_data.get("role") != "admin":
            return JSONResponse(status_code=403, content={
                "success": False, "detail": "Admin access required"
            })
        
        data = await request.json()
        
        if not data.get('username') or not data.get('password'):
            return JSONResponse(status_code=400, content={
                "success": False, "detail": "Username and password required"
            })
        
        with Session(dqa_engine) as session:
            existing = session.query(User).filter(User.username == data['username']).first()
            if existing:
                return JSONResponse(status_code=400, content={
                    "success": False, "detail": "Username already exists"
                })
            
            new_user = User(
                full_name=data.get('full_name', data['username']),
                username=data['username'],
                password_hash=hash_password(data['password']),
                role=data.get('role', 'dqa_team'),
                position=data.get('position', ''),
                created_by=session_data['username']
            )
            session.add(new_user)
            session.commit()
            
            logger.info(f"✅ User created: {data['username']} by {session_data['username']}")
            
            return {
                "success": True,
                "message": f"User '{data['username']}' created successfully"
            }
            
    except Exception as e:
        logger.error(f"Create user error: {e}")
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})

@app.put("/api/users/{user_id}")
async def update_user(user_id: int, request: Request):
    """Update user details including password (Admin only)"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import User
        from sqlalchemy.orm import Session
        
        token = request.headers.get("X-Session-Token", "")
        session_data = get_current_user_from_token(token)
        
        if not session_data or session_data.get("role") != "admin":
            return JSONResponse(status_code=403, content={
                "success": False, "detail": "Admin access required"
            })
        
        data = await request.json()
        
        with Session(dqa_engine) as session:
            user = session.query(User).filter(User.id == user_id).first()
            if not user:
                return JSONResponse(status_code=404, content={
                    "success": False, "detail": "User not found"
                })
            
            # Update fields if provided
            if 'full_name' in data and data['full_name']:
                user.full_name = data['full_name']
            if 'role' in data and data['role']:
                user.role = data['role']
            if 'position' in data:
                user.position = data['position']
            if 'is_active' in data:
                user.is_active = data['is_active']
            if 'password' in data and data['password']:
                user.password_hash = hash_password(data['password'])
            
            session.commit()
            
            logger.info(f"✅ User updated: {user.username} by {session_data['username']}")
            
            return {"success": True, "message": f"User '{user.username}' updated successfully"}
            
    except Exception as e:
        logger.error(f"Update user error: {e}")
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})


@app.delete("/api/users/{user_id}")
async def delete_user(user_id: int, request: Request):
    """Delete a user (Admin only)"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import User
        from sqlalchemy.orm import Session
        
        token = request.headers.get("X-Session-Token", "")
        session_data = get_current_user_from_token(token)
        
        if not session_data or session_data.get("role") != "admin":
            return JSONResponse(status_code=403, content={
                "success": False, "detail": "Admin access required"
            })
        
        with Session(dqa_engine) as session:
            user = session.query(User).filter(User.id == user_id).first()
            if not user:
                return JSONResponse(status_code=404, content={
                    "success": False, "detail": "User not found"
                })
            
            # Don't allow deleting yourself
            if user.username == session_data['username']:
                return JSONResponse(status_code=400, content={
                    "success": False, "detail": "You cannot delete your own account"
                })
            
            username = user.username
            session.delete(user)
            session.commit()
            
            logger.info(f"🗑️ User deleted: {username} by {session_data['username']}")
            
            return {"success": True, "message": f"User '{username}' deleted successfully"}
            
    except Exception as e:
        logger.error(f"Delete user error: {e}")
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})    
# ============================================================================
# STARTUP EVENT
# ============================================================================

@app.on_event("shutdown")
async def shutdown_event():
    """Cleanup on shutdown"""
    logger.info("MedDQA shutting down...")
    record_locks.clear()
    active_sessions.clear()

@app.post("/api/review/complete")
async def complete_review_workflow(request: Request):
    """
    Complete the review workflow and save results to proper columns.
    Extracts username from session token for reliable attribution.
    """
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import CareCardRecord, DQAAuditLog, User
        from sqlalchemy.orm import Session
        import json, hashlib
        from datetime import datetime
        
        body = await request.json()
        
        hospital_number = body.get("hospital_number")
        person_uuid = body.get("person_uuid")
        review_results = body.get("review_results", [])
        total_fields = body.get("total_fields", 0)
        matched_fields = body.get("matched_fields", 0)
        corrected_on_emr = body.get("corrected_on_emr", 0)
        corrected_on_carecard = body.get("corrected_on_carecard", 0)
        
        # ================================================================
        # GET USERNAME
        # ================================================================
        verified_by = body.get("verified_by", None)
        if not verified_by or verified_by == "Unknown":
            verified_by = request.headers.get("X-User", None)
        if not verified_by or verified_by == "Unknown":
            verified_by = request.headers.get("x-user", "Unknown")
        
        logger.info(f"📝 Saving review for {hospital_number} by: {verified_by}")
        
        affected_visits_map = body.get("affected_visits", {})  # {field: [indices]}

        # ================================================================
        # SEPARATE RESULTS BY TYPE
        # ================================================================
        biodata_results = [r for r in review_results if r.get("step") == 1]
        latest_refill_results = [r for r in review_results if r.get("step") == 2]
        refill_batch_results = [r for r in review_results if r.get("step") == 3]
        vl_batch_results = [r for r in review_results if r.get("step") == 4]
        
        # ================================================================
        # BUILD CLEAN BIODATA WITH ALL FIELDS PRESERVED
        # ================================================================
        biodata_clean = []
        for r in biodata_results:
            biodata_clean.append({
                "field": r.get("field"),
                "label": r.get("label"),
                "emr_value": r.get("emr_value"),
                "match": r.get("match", False),
                "corrected_on": r.get("corrected_on"),
                "discrepancy_type": r.get("discrepancy_type"),
                "discrepancy_note": r.get("discrepancy_note"),
                "care_card_value": r.get("care_card_value"),
                "original_emr_value": r.get("original_emr_value"),
                "affected_visits": r.get("affected_visits") or affected_visits_map.get(r.get("field"), []),
                "step": 1
            })
        
        # ================================================================
        # BUILD CLEAN LATEST REFILL WITH ALL FIELDS PRESERVED
        # ================================================================
        latest_refill_clean = []
        for r in latest_refill_results:
            latest_refill_clean.append({
                "field": r.get("field"),
                "label": r.get("label"),
                "emr_value": r.get("emr_value"),
                "match": r.get("match", False),
                "corrected_on": r.get("corrected_on"),
                "discrepancy_type": r.get("discrepancy_type"),
                "discrepancy_note": r.get("discrepancy_note"),
                "care_card_value": r.get("care_card_value"),
                "original_emr_value": r.get("original_emr_value"),
                "affected_visits": r.get("affected_visits") or affected_visits_map.get(r.get("field"), []),
                "step": 2
            })
        
        # ================================================================
        # BUILD ENROLLMENT DATA
        # ================================================================
        enrollment_data = {
            "biodata_verification": biodata_clean,
            "biodata_matched": sum(1 for r in biodata_clean if r.get("match")),
            "biodata_total": len(biodata_clean),
            "latest_refill_verification": latest_refill_clean,
            "latest_refill_matched": sum(1 for r in latest_refill_clean if r.get("match")),
            "latest_refill_total": len(latest_refill_clean),
            "review_completed": True,
            "verified_by": verified_by,
            "verified_at": datetime.utcnow().isoformat()
        }
        
        # ================================================================
        # BUILD DRUG PICKUPS WITH ALL FIELDS INCLUDING DISCREPANCY INFO
        # ================================================================
        drug_pickups = []
        for r in refill_batch_results:
            drug_pickups.append({
                "field": r.get("field"),
                "label": r.get("label"),
                "emr_summary": r.get("emr_value"),
                "match": r.get("match", False),
                "corrected_on": r.get("corrected_on"),
                "discrepancy_type": r.get("discrepancy_type"),
                "discrepancy_note": r.get("discrepancy_note"),
                "care_card_value": r.get("care_card_value"),
                "original_emr_value": r.get("original_emr_value"),
                "affected_visits": r.get("affected_visits") or affected_visits_map.get(r.get("field"), []),
                "reviewed_by": verified_by,
                "reviewed_at": datetime.utcnow().isoformat()
            })
        
        # ================================================================
        # BUILD VIRAL LOADS WITH ALL FIELDS INCLUDING DISCREPANCY INFO
        # ================================================================
        viral_loads = []
        for r in vl_batch_results:
            viral_loads.append({
                "field": r.get("field"),
                "label": r.get("label"),
                "emr_summary": r.get("emr_value"),
                "match": r.get("match", False),
                "corrected_on": r.get("corrected_on"),
                "discrepancy_type": r.get("discrepancy_type"),
                "discrepancy_note": r.get("discrepancy_note"),
                "care_card_value": r.get("care_card_value"),
                "original_emr_value": r.get("original_emr_value"),
                "affected_visits": r.get("affected_visits") or affected_visits_map.get(r.get("field"), []),
                "reviewed_by": verified_by,
                "reviewed_at": datetime.utcnow().isoformat()
            })
        
        # ================================================================
        # CALCULATE STATISTICS
        # ================================================================
        match_rate = f"{round((matched_fields / total_fields * 100), 1)}%" if total_fields > 0 else "0%"
        mismatched_fields = total_fields - matched_fields
        now = datetime.utcnow()
        
        # ================================================================
        # SAVE TO DATABASE
        # ================================================================
        with Session(dqa_engine) as session:
            existing = session.query(CareCardRecord).filter(
                CareCardRecord.hospital_number == hospital_number
            ).first()
            
            if existing:
                existing.enrollment_data = enrollment_data
                if drug_pickups:
                    existing.drug_pickups = drug_pickups
                if viral_loads:
                    existing.viral_loads = viral_loads
                existing.is_verified = True
                existing.verified_by = verified_by
                existing.verified_at = now
                existing.updated_by = verified_by
                existing.created_by = existing.created_by or verified_by
                existing.updated_at = now
                logger.info(f"📝 Updated care card record for {hospital_number}")
            else:
                new_record = CareCardRecord(
                    hospital_number=hospital_number,
                    person_uuid=person_uuid,
                    enrollment_data=enrollment_data,
                    drug_pickups=drug_pickups if drug_pickups else [],
                    viral_loads=viral_loads if viral_loads else [],
                    is_verified=True,
                    verified_by=verified_by,
                    verified_at=now,
                    created_by=verified_by,
                    updated_by=verified_by
                )
                session.add(new_record)
                logger.info(f"📝 Created new care card record for {hospital_number}")
            
            # ================================================================
            # SAVE AUDIT LOG
            # ================================================================
            audit_log = DQAAuditLog(
                hospital_number=hospital_number,
                person_uuid=person_uuid,
                care_card_data={
                    "biodata_results": biodata_clean,
                    "refill_results": refill_batch_results,
                    "vl_results": vl_batch_results,
                    "total_fields": total_fields,
                    "matched_fields": matched_fields,
                    "mismatched_fields": mismatched_fields,
                    "corrected_on_emr": corrected_on_emr,
                    "corrected_on_carecard": corrected_on_carecard,
                    "match_rate": match_rate
                },
                validation_status="Matched" if mismatched_fields == 0 else "Corrected",
                discrepancies_found=(mismatched_fields > 0),
                issues_fixed=corrected_on_emr + corrected_on_carecard,
                total_comparisons=total_fields,
                matched_comparisons=matched_fields,
                user_name=verified_by
            )
            session.add(audit_log)
            session.commit()
            
            logger.info(f"✅ Review saved - Audit ID: {audit_log.id}, By: {verified_by}")
            
            # ✅ LOG WHAT WAS SAVED FOR DEBUGGING
            logger.info(f"   Biodata items: {len(biodata_clean)}")
            logger.info(f"   Latest refill items: {len(latest_refill_clean)}")
            logger.info(f"   Drug pickups: {len(drug_pickups)}")
            logger.info(f"   Viral loads: {len(viral_loads)}")
            
            # Log discrepancy details
            for item in biodata_clean + latest_refill_clean:
                if not item.get("match") and item.get("discrepancy_type"):
                    logger.info(f"   Discrepancy: {item['field']} -> {item['discrepancy_type']} | Note: {item.get('discrepancy_note', 'N/A')}")
            
            for item in drug_pickups + viral_loads:
                if not item.get("match") and item.get("discrepancy_type"):
                    logger.info(f"   Batch Discrepancy: {item['field']} -> {item['discrepancy_type']} | Note: {item.get('discrepancy_note', 'N/A')}")
        
        # ================================================================
        # RETURN RESPONSE
        # ================================================================
        return {
            "success": True,
            "audit_id": audit_log.id,
            "summary": {
                "total": total_fields,
                "matched": matched_fields,
                "mismatched": mismatched_fields,
                "match_rate": match_rate,
                "corrected_emr": corrected_on_emr,
                "corrected_carecard": corrected_on_carecard,
                "verified_by": verified_by,
                "breakdown": {
                    "biodata": f"{sum(1 for r in biodata_clean if r.get('match'))}/{len(biodata_clean)}",
                    "latest_refill": f"{sum(1 for r in latest_refill_clean if r.get('match'))}/{len(latest_refill_clean)}",
                    "refill_history": f"{sum(1 for r in refill_batch_results if r.get('match'))}/{len(refill_batch_results)}",
                    "vl_history": f"{sum(1 for r in vl_batch_results if r.get('match'))}/{len(vl_batch_results)}"
                }
            }
        }
        
    except Exception as e:
        logger.error(f"Review completion error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})
    

@app.get("/api/review/start/{hospital_number:path}")
async def start_review_workflow(hospital_number: str, request: Request):
    """
    Start the step-by-step care card review workflow.
    Returns patient EMR data AND any previously saved review.
    Loads from the correct columns: enrollment_data, drug_pickups, viral_loads
    """
    try:
        from app.database import emr_engine, dqa_engine
        from app.models.dqa_models import CareCardRecord
        from sqlalchemy.orm import Session
        
        hospital_number = unquote(hospital_number)
        logger.info(f"📋 Starting review workflow for: {hospital_number}")
        
        # ================================================================
        # LOAD PREVIOUS REVIEW FROM CORRECT COLUMNS
        # ================================================================
# In previous_review:
        previous_review = {
            "biodata": [],
            "latest_refill": [],
            "refill_batch": {},
            "vl_batch": {},
            "care_card_values": {},
            "original_emr_values": {},
            "discrepancy_types": {},     
            "discrepancy_notes": {},     
            "is_complete": False
        }
        
        drug_pickups_details = []
        viral_loads_details = []
        care_record = None  # ✅ Initialize before try block
        
        try:
            with Session(dqa_engine) as session:
                care_record = session.query(CareCardRecord).filter(
                    CareCardRecord.hospital_number == hospital_number
                ).first()
                
                if care_record:
                    logger.info(f"📂 Found care card record for {hospital_number}")
                    
                    # Load biodata + latest refill from enrollment_data
                    if care_record.enrollment_data and isinstance(care_record.enrollment_data, dict):
                        ed = care_record.enrollment_data
                        previous_review["biodata"] = ed.get("biodata_verification", [])
                        previous_review["latest_refill"] = ed.get("latest_refill_verification", [])
                        previous_review["is_complete"] = ed.get("review_completed", False)
                        
                        # Load direct maps from enrollment_data
                        previous_review["discrepancy_types"] = ed.get("discrepancy_types", {})
                        previous_review["discrepancy_notes"] = ed.get("discrepancy_notes", {})
                        
                        # Extract care_card_values and original_emr_values from biodata
                        for item in previous_review["biodata"]:
                            if isinstance(item, dict):
                                fld = item.get("field")
                                if fld:
                                    if item.get("care_card_value"):
                                        previous_review["care_card_values"][fld] = item["care_card_value"]
                                    if item.get("original_emr_value"):
                                        previous_review["original_emr_values"][fld] = item["original_emr_value"]
                                    if item.get("discrepancy_type"):
                                        previous_review["discrepancy_types"][fld] = item["discrepancy_type"]
                                    if item.get("discrepancy_note"):
                                        previous_review["discrepancy_notes"][fld] = item["discrepancy_note"]
                        
                        # Extract from latest_refill
                        for item in previous_review["latest_refill"]:
                            if isinstance(item, dict):
                                fld = item.get("field")
                                if fld:
                                    if item.get("care_card_value"):
                                        previous_review["care_card_values"][fld] = item["care_card_value"]
                                    if item.get("original_emr_value"):
                                        previous_review["original_emr_values"][fld] = item["original_emr_value"]
                                    if item.get("discrepancy_type"):
                                        previous_review["discrepancy_types"][fld] = item["discrepancy_type"]
                                    if item.get("discrepancy_note"):
                                        previous_review["discrepancy_notes"][fld] = item["discrepancy_note"]
                    
                    # Load refill batch results from drug_pickups
                    if care_record.drug_pickups and isinstance(care_record.drug_pickups, list):
                        for item in care_record.drug_pickups:
                            if isinstance(item, dict) and item.get("field"):
                                field = item.get("field")
                                corrected_on = item.get("corrected_on")
                                
                                if item.get("match"):
                                    previous_review["refill_batch"][field] = "match"
                                elif corrected_on:
                                    previous_review["refill_batch"][field] = f"corrected_{corrected_on}"
                                
                                # Build drug_pickups_details for frontend
                                drug_pickups_details.append({
                                    "field": field,
                                    "care_card_value": item.get("care_card_value"),
                                    "original_emr_value": item.get("original_emr_value"),
                                    "discrepancy_type": item.get("discrepancy_type"),
                                    "discrepancy_note": item.get("discrepancy_note"),
                                    "match": item.get("match", False),
                                    "corrected_on": corrected_on
                                })
                                
                                # Extract values to maps
                                if item.get("care_card_value"):
                                    previous_review["care_card_values"][field] = item["care_card_value"]
                                if item.get("discrepancy_type"):
                                    previous_review["discrepancy_types"][field] = item["discrepancy_type"]
                                if item.get("discrepancy_note"):
                                    previous_review["discrepancy_notes"][field] = item["discrepancy_note"]
                    
                    # Load VL batch results from viral_loads
                    if care_record.viral_loads and isinstance(care_record.viral_loads, list):
                        for item in care_record.viral_loads:
                            if isinstance(item, dict) and item.get("field"):
                                field = item.get("field")
                                corrected_on = item.get("corrected_on")
                                
                                if item.get("match"):
                                    previous_review["vl_batch"][field] = "match"
                                elif corrected_on:
                                    previous_review["vl_batch"][field] = f"corrected_{corrected_on}"
                                
                                # Build viral_loads_details for frontend
                                viral_loads_details.append({
                                    "field": field,
                                    "care_card_value": item.get("care_card_value"),
                                    "original_emr_value": item.get("original_emr_value"),
                                    "discrepancy_type": item.get("discrepancy_type"),
                                    "discrepancy_note": item.get("discrepancy_note"),
                                    "match": item.get("match", False),
                                    "corrected_on": corrected_on
                                })
                                
                                # Extract values to maps
                                if item.get("care_card_value"):
                                    previous_review["care_card_values"][field] = item["care_card_value"]
                                if item.get("discrepancy_type"):
                                    previous_review["discrepancy_types"][field] = item["discrepancy_type"]
                                if item.get("discrepancy_note"):
                                    previous_review["discrepancy_notes"][field] = item["discrepancy_note"]
                    
                    logger.info(f"✅ Loaded: biodata={len(previous_review['biodata'])}, refill_batch={len(previous_review['refill_batch'])}, vl_batch={len(previous_review['vl_batch'])}")
                    logger.info(f"✅ drug_pickups_details={len(drug_pickups_details)}, viral_loads_details={len(viral_loads_details)}")
                    
        except Exception as e:
            logger.warning(f"Could not load previous review: {e}")
            import traceback
            traceback.print_exc()
        
        # ✅ Now safe to use care_record (might be None) after the try block
        # The direct maps from enrollment_data are already loaded above
        
        # If care_record was loaded, also extract from enrollment_data direct maps
        # (This is redundant with the loop above but ensures coverage)
        if care_record and care_record.enrollment_data and isinstance(care_record.enrollment_data, dict):
            ed = care_record.enrollment_data
            # Merge direct maps (loop values take priority since they're more specific)
            direct_types = ed.get("discrepancy_types", {})
            direct_notes = ed.get("discrepancy_notes", {})
            if direct_types:
                for k, v in direct_types.items():
                    if k not in previous_review["discrepancy_types"]:
                        previous_review["discrepancy_types"][k] = v
            if direct_notes:
                for k, v in direct_notes.items():
                    if k not in previous_review["discrepancy_notes"]:
                        previous_review["discrepancy_notes"][k] = v
        
        # ================================================================
        # GET EMR DATA
        # ================================================================
        with emr_engine.connect() as conn:
            # Get patient bio
            patient = conn.execute(
                text("""
                    SELECT 
                        p.first_name, p.surname, p.other_name,
                        INITCAP(p.sex) AS sex,
                        p.date_of_birth,
                        hac.visit_date::DATE AS art_start_date,
                        facility.name AS facility_name
                    FROM patient_person p
                    INNER JOIN hiv_enrollment h ON h.person_uuid = p.uuid AND h.archived = 0
                    INNER JOIN base_organisation_unit facility ON facility.id = h.facility_id
                    LEFT JOIN hiv_art_clinical hac ON hac.hiv_enrollment_uuid = h.uuid 
                        AND hac.archived = 0 AND hac.is_commencement = TRUE
                    WHERE p.hospital_number = :hn AND p.archived = 0 LIMIT 1
                """),
                {"hn": hospital_number}
            ).fetchone()
            
            if not patient:
                return JSONResponse(status_code=404, content={"success": False, "detail": "Patient not found"})
            
            person_result = conn.execute(
                text("SELECT uuid FROM patient_person WHERE hospital_number = :hn AND archived = 0"),
                {"hn": hospital_number}
            ).fetchone()
            person_uuid = person_result[0] if person_result else None
            
            # Get refills
            refills = []
            if person_uuid:
                refill_rows = conn.execute(
                    text("""
                        SELECT 
                            hap.visit_date AS pickup_date,
                            COALESCE((elem.value ->> 'duration')::INTEGER, hap.refill_period, 0) AS duration,
                            elem.value ->> 'regimenName' AS regimen_name,
                            hap.next_appointment
                        FROM hiv_art_pharmacy hap
                        CROSS JOIN LATERAL jsonb_array_elements(hap.extra -> 'regimens') WITH ORDINALITY AS elem(value, ordinality)
                        WHERE hap.person_uuid = :uuid AND hap.archived = 0 AND elem.ordinality = 1
                        ORDER BY hap.visit_date DESC
                    """),
                    {"uuid": person_uuid}
                ).fetchall()
                refills = [{"date": str(r[0])[:10] if r[0] else 'N/A', "duration": f"{int(r[1] or 0)} days", "regimen": r[2] or 'N/A', "next_appt": str(r[3])[:10] if r[3] else 'N/A'} for r in refill_rows]
            
            # Get VL results
            viral_loads = []
            if person_uuid:
                vl_rows = conn.execute(
                    text("""
                        SELECT CAST(ls.date_sample_collected AS DATE) AS sample_date, sm.result_reported AS vl_result, CAST(sm.date_result_reported AS DATE) AS result_date
                        FROM laboratory_result sm INNER JOIN laboratory_test lt ON lt.id = sm.test_id INNER JOIN laboratory_sample ls ON ls.test_id = lt.id
                        WHERE lt.lab_test_id = 16 AND sm.patient_uuid = :uuid AND sm.result_reported IS NOT NULL AND sm.archived = 0
                        ORDER BY ls.date_sample_collected DESC
                    """),
                    {"uuid": person_uuid}
                ).fetchall()
                viral_loads = [{"sample_date": str(v[0])[:10] if v[0] else 'N/A', "result": v[1] or 'N/A', "result_date": str(v[2])[:10] if v[2] else 'N/A'} for v in vl_rows]
            
            latest_refill = refills[0] if refills else None
            
            workflow = {
                "patient_name": f"{patient[0]} {patient[1]}",
                "hospital_number": hospital_number,
                "previous_review": previous_review,
                "drug_pickups_details": drug_pickups_details,  # ✅ ADD THIS
                "viral_loads_details": viral_loads_details,    # ✅ ADD THIS
                "steps": [
                    {
                        "step": 1,
                        "title": "Biodata Verification",
                        "icon": "bi-person-badge",
                        "instruction": "Verify each biodata field against the physical Care Card.",
                        "fields": [
                            {"field": "sex", "label": "Sex", "emr_value": patient[3] or 'N/A'},
                            {"field": "date_of_birth", "label": "Date of Birth", "emr_value": str(patient[4])[:10] if patient[4] else 'N/A'},
                            {"field": "art_start_date", "label": "ART Start Date", "emr_value": str(patient[5])[:10] if patient[5] else 'N/A'}
                        ]
                    },
                    {
                        "step": 2,
                        "title": "Latest Refill Details",
                        "icon": "bi-capsule",
                        "instruction": "Verify the most recent refill details against the Care Card.",
                        "fields": [
                            {"field": "last_pickup_date", "label": "Last Pickup Date", "emr_value": latest_refill['date'] if latest_refill else 'N/A'},
                            {"field": "refill_duration", "label": "Refill Duration", "emr_value": latest_refill['duration'] if latest_refill else 'N/A'},
                            {"field": "regimen", "label": "Current Regimen", "emr_value": latest_refill['regimen'] if latest_refill else 'N/A'}
                        ]
                    },
                    {
                        "step": 3,
                        "title": "Refill History Review",
                        "icon": "bi-calendar3",
                        "instruction": "Review all refill records as a batch. Select overall status for each category.",
                        "refills": refills
                    },
                    {
                        "step": 4,
                        "title": "Viral Load History",
                        "icon": "bi-droplet",
                        "instruction": "Review all viral load results as a batch. Select overall status for each category.",
                        "viral_loads": viral_loads
                    }
                ]
            }
            
            logger.info(f"✅ Returning workflow with drug_pickups_details={len(drug_pickups_details)}, viral_loads_details={len(viral_loads_details)}")
            
            return {"success": True, "workflow": workflow}
            
    except Exception as e:
        logger.error(f"Review workflow error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})
    
# ============================================================================
# DQA PERFORMANCE DASHBOARD API
# ============================================================================

@app.get("/api/dashboard/performance")
async def get_performance_data(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    facility: str = None
):
    """
    Returns performance data for all assessors.
    Includes leaderboard rankings, stats, and weekly trend data.
    """
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import CareCardRecord, User
        from sqlalchemy.orm import Session
        from sqlalchemy import func, text
        from datetime import datetime, timedelta
        
        with Session(dqa_engine) as session:
            # Base query for verified records
            query = session.query(CareCardRecord).filter(CareCardRecord.is_verified == True)
            
            if start_date:
                query = query.filter(CareCardRecord.verified_at >= start_date)
            if end_date:
                query = query.filter(CareCardRecord.verified_at <= end_date + "T23:59:59")
            
            # Facility filter
            if facility:
                from app.database import emr_engine
                # Get hospital numbers for this facility
                with emr_engine.connect() as conn:
                    facility_hns = conn.execute(text("""
                        SELECT p.hospital_number
                        FROM patient_person p
                        INNER JOIN hiv_enrollment h ON h.person_uuid = p.uuid AND h.archived = 0
                        INNER JOIN base_organisation_unit f ON f.id = h.facility_id
                        WHERE f.name = :facility AND p.archived = 0
                    """), {"facility": facility}).fetchall()
                    hn_list = [r[0] for r in facility_hns]
                    if hn_list:
                        query = query.filter(CareCardRecord.hospital_number.in_(hn_list))
                    else:
                        query = query.filter(CareCardRecord.hospital_number == "__none__")
            
            records = query.all()
            
            # Get all users
            users = session.query(User).filter(User.is_active == True).all()
            
            # Build assessor performance map
            assessor_map = {}
            for user in users:
                username = user.full_name or user.username
                assessor_map[username] = {
                    "username": user.username,
                    "full_name": user.full_name,
                    "role": user.role,
                    "position": user.position or "",
                    "total_assessed": 0,
                    "total_fields": 0,
                    "matched": 0,
                    "corrected_emr": 0,
                    "corrected_carecard": 0,
                    "last_active": str(user.last_login) if user.last_login else None
                }
            
            # Aggregate record data
            for record in records:
                verified_by = record.verified_by or record.created_by or "Unknown"
                
                if verified_by not in assessor_map:
                    assessor_map[verified_by] = {
                        "username": verified_by,
                        "full_name": verified_by,
                        "role": "dqa_team",
                        "position": "",
                        "total_assessed": 0,
                        "total_fields": 0,
                        "matched": 0,
                        "corrected_emr": 0,
                        "corrected_carecard": 0,
                        "last_active": str(record.verified_at) if record.verified_at else None
                    }
                
                ed = record.enrollment_data or {}
                assessor_map[verified_by]["total_assessed"] += 1
                assessor_map[verified_by]["total_fields"] += ed.get("biodata_total", 0) + ed.get("latest_refill_total", 0)
                assessor_map[verified_by]["matched"] += ed.get("biodata_matched", 0) + ed.get("latest_refill_matched", 0)
                
                # Count batch corrections
                for dp in (record.drug_pickups or []):
                    if dp.get("corrected_on") == "emr":
                        assessor_map[verified_by]["corrected_emr"] += 1
                    elif dp.get("corrected_on") == "care_card":
                        assessor_map[verified_by]["corrected_carecard"] += 1
                
                for vl in (record.viral_loads or []):
                    if vl.get("corrected_on") == "emr":
                        assessor_map[verified_by]["corrected_emr"] += 1
                    elif vl.get("corrected_on") == "care_card":
                        assessor_map[verified_by]["corrected_carecard"] += 1
                
                # Update last active
                if record.verified_at:
                    current_last = assessor_map[verified_by].get("last_active")
                    if not current_last or str(record.verified_at) > current_last:
                        assessor_map[verified_by]["last_active"] = str(record.verified_at)
            
            # Sort by total_assessed descending
            assessors = sorted(assessor_map.values(), key=lambda x: x["total_assessed"], reverse=True)
            
            # Calculate stats
            total_assessed = sum(a["total_assessed"] for a in assessors)
            active_assessors = sum(1 for a in assessors if a["total_assessed"] > 0)
            top_performer = assessors[0]["full_name"] if assessors else "—"
            
            total_fields_all = sum(a["total_fields"] for a in assessors)
            total_matched_all = sum(a["matched"] for a in assessors)
            overall_match_rate = f"{round((total_matched_all / total_fields_all * 100), 1)}%" if total_fields_all > 0 else "0%"
            
            # Build weekly trend data
            trend_labels = []
            trend_assessors = {}
            
            if records:
                # Last 8 weeks
                now = datetime.utcnow()
                for week_offset in range(7, -1, -1):
                    week_start = (now - timedelta(days=week_offset * 7)).strftime('%Y-%m-%d')
                    week_end = (now - timedelta(days=(week_offset - 1) * 7)).strftime('%Y-%m-%d')
                    trend_labels.append(f"{week_start[5:]} to {week_end[5:]}")
                    
                    for assessor in assessors:
                        name = assessor["full_name"]
                        if name not in trend_assessors:
                            trend_assessors[name] = []
                        
                        count = session.query(CareCardRecord).filter(
                            CareCardRecord.is_verified == True,
                            func.date(CareCardRecord.verified_at) >= week_start,
                            func.date(CareCardRecord.verified_at) < week_end,
                            (CareCardRecord.verified_by == name) | (CareCardRecord.created_by == name)
                        ).count()
                        trend_assessors[name].append(count)
            
            return {
                "success": True,
                "total_assessed": total_assessed,
                "active_assessors": active_assessors,
                "top_performer": top_performer,
                "overall_match_rate": overall_match_rate,
                "assessors": assessors,
                "trend": {
                    "labels": trend_labels,
                    "assessors": trend_assessors
                }
            }
            
    except Exception as e:
        logger.error(f"Dashboard error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})


@app.get("/api/dashboard/facilities")
async def get_facilities(request: Request):
    """Get list of facilities for filter dropdown"""
    try:
        from app.database import emr_engine
        from sqlalchemy import text
        
        with emr_engine.connect() as conn:
            result = conn.execute(text("""
                SELECT DISTINCT facility.name
                FROM patient_person p
                INNER JOIN hiv_enrollment h ON h.person_uuid = p.uuid AND h.archived = 0
                INNER JOIN base_organisation_unit facility ON facility.id = h.facility_id
                WHERE p.archived = 0
                ORDER BY facility.name
            """)).fetchall()
            facilities = [r[0] for r in result if r[0]]
        
        return {"success": True, "facilities": facilities}
    except Exception as e:
        return {"success": True, "facilities": []}


@app.get("/api/dashboard/performance/excel")
async def export_performance_excel(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    facility: str = None
):
    """Export performance data to Excel"""
    try:
        from app.database import dqa_engine, emr_engine
        from app.models.dqa_models import CareCardRecord, User
        from sqlalchemy.orm import Session
        from sqlalchemy import text
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        import io
        
        with Session(dqa_engine) as session:
            query = session.query(CareCardRecord).filter(CareCardRecord.is_verified == True)
            if start_date:
                query = query.filter(CareCardRecord.verified_at >= start_date)
            if end_date:
                query = query.filter(CareCardRecord.verified_at <= end_date + "T23:59:59")
            
            records = query.all()
        
        # Aggregate by assessor
        assessor_map = {}
        for record in records:
            verified_by = record.verified_by or record.created_by or "Unknown"
            if verified_by not in assessor_map:
                assessor_map[verified_by] = {"name": verified_by, "folders": 0, "matched": 0, "total_fields": 0, "corrected_emr": 0, "corrected_carecard": 0}
            
            ed = record.enrollment_data or {}
            assessor_map[verified_by]["folders"] += 1
            assessor_map[verified_by]["total_fields"] += ed.get("biodata_total", 0) + ed.get("latest_refill_total", 0)
            assessor_map[verified_by]["matched"] += ed.get("biodata_matched", 0) + ed.get("latest_refill_matched", 0)
            
            for dp in (record.drug_pickups or []):
                if dp.get("corrected_on") == "emr": assessor_map[verified_by]["corrected_emr"] += 1
                elif dp.get("corrected_on") == "care_card": assessor_map[verified_by]["corrected_carecard"] += 1
            for vl in (record.viral_loads or []):
                if vl.get("corrected_on") == "emr": assessor_map[verified_by]["corrected_emr"] += 1
                elif vl.get("corrected_on") == "care_card": assessor_map[verified_by]["corrected_carecard"] += 1
        
        assessors = sorted(assessor_map.values(), key=lambda x: x["folders"], reverse=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DQA Performance"
        
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        thin_border = Border(left=Side(style='thin', color='d1d5db'), right=Side(style='thin', color='d1d5db'),
                             top=Side(style='thin', color='d1d5db'), bottom=Side(style='thin', color='d1d5db'))
        
        headers = ["Rank", "Assessor Name", "Folders Assessed", "Total Fields", "Matched", "Match Rate", "Corrected (EMR)", "Corrected (Care Card)"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = thin_border
        ws.row_dimensions[1].height = 32
        
        for i, a in enumerate(assessors):
            row = i + 2
            match_rate = f"{round((a['matched']/a['total_fields']*100),1)}%" if a['total_fields'] > 0 else "0%"
            vals = [i+1, a['name'], a['folders'], a['total_fields'], a['matched'], match_rate, a['corrected_emr'], a['corrected_carecard']]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.border = thin_border; c.font = Font(size=9, name="Calibri")
                c.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
        
        widths = {1:6, 2:30, 3:16, 4:14, 5:12, 6:12, 7:16, 8:18}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = 'A2'
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename=DQA_Performance_{datetime.now().strftime('%Y%m%d')}.xlsx"})
    except Exception as e:
        logger.error(f"Export error: {e}")
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})
    
@app.get("/dashboard")
async def serve_dashboard():
    """Serve the performance dashboard page"""
    from fastapi.responses import FileResponse
    import os
    
    # Get the base directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Try multiple possible locations
    possible_paths = [
        os.path.join(base_dir, "templates", "dashboard.html"),
        os.path.join(base_dir, "..", "templates", "dashboard.html"),
        os.path.join(base_dir, "..", "static", "dashboard.html"),
        os.path.join(base_dir, "static", "dashboard.html"),
    ]
    
    # If frozen (PyInstaller)
    if getattr(sys, 'frozen', False):
        possible_paths.insert(0, os.path.join(sys._MEIPASS, "app", "templates", "dashboard.html"))
        possible_paths.insert(0, os.path.join(sys._MEIPASS, "templates", "dashboard.html"))
    
    for path in possible_paths:
        abs_path = os.path.abspath(path)
        logger.info(f"Checking path: {abs_path}")
        if os.path.exists(abs_path):
            logger.info(f"✅ Found dashboard at: {abs_path}")
            return FileResponse(abs_path, media_type="text/html")
    
    logger.error(f"❌ Dashboard not found in any of: {possible_paths}")
    return JSONResponse(
        status_code=404,
        content={"error": f"Dashboard not found. Checked: {possible_paths}"}
    )

from app.routers.singlereport import generate_dqa_verification_pdf, get_status, DISC_LABELS

@app.get("/api/reports/dqa-verification/{hospital_number:path}")
async def dqa_verification_pdf_route(hospital_number: str, request: Request):
    return await generate_dqa_verification_pdf(hospital_number, request)
# ============================================================================
# BATCH PDF - NO COVER PAGE, WIDE TABLE, PROFESSIONAL
# ============================================================================
from app.routers.batchreport import generate_batch_verification_pdf

@app.get("/api/reports/dqa-verification-batch")
async def dqa_verification_batch_route(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    username: str = None
):
    return await generate_batch_verification_pdf(
        request,
        start_date,
        end_date,
        username
    )
# ============================================================================
# BATCH COUNT
# ============================================================================

@app.get("/api/reports/batch-count")
async def get_batch_count(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    username: str = None
):
    """Get count of verified records for batch print preview"""
    try:
        from app.database import dqa_engine
        from app.models.dqa_models import CareCardRecord
        from sqlalchemy.orm import Session
        from sqlalchemy import cast, Date
        
        with Session(dqa_engine) as session:
            query = session.query(CareCardRecord).filter(CareCardRecord.is_verified == True)
            
            if username:
                query = query.filter(
                    (CareCardRecord.verified_by == username) | 
                    (CareCardRecord.created_by == username)
                )
            
            if start_date:
                query = query.filter(cast(CareCardRecord.verified_at, Date) >= start_date)
            if end_date:
                query = query.filter(cast(CareCardRecord.verified_at, Date) <= end_date)
            
            count = query.count()
        
        return {"success": True, "count": count}
    except Exception as e:
        return {"success": False, "count": 0, "detail": str(e)}

@app.put("/api/patients/update-vl")
async def update_viral_load(request: Request):
    """
    Update Viral Load laboratory data in EMR.
    Updates the exact VL record matching the sample collection date.
    """
    try:
        from app.database import emr_engine
        from sqlalchemy import text
        
        body = await request.json()
        hospital_number = body.get("hospital_number")
        field_name = body.get("field_name")
        new_value = body.get("new_value")
        sample_date = body.get("sample_date")
        user = request.headers.get("X-User", "Unknown")
        
        if not hospital_number or not field_name or not new_value:
            return JSONResponse(status_code=400, content={"success": False, "detail": "Missing required fields"})
        
        logger.info(f"🔬 Updating VL {field_name} for {hospital_number} (sample: {sample_date}) to '{new_value}' by {user}")
        
        with emr_engine.connect() as conn:
            # Get person_uuid
            person = conn.execute(text(
                "SELECT uuid FROM patient_person WHERE hospital_number = :hn AND archived = 0"
            ), {"hn": hospital_number}).fetchone()
            
            if not person:
                return JSONResponse(status_code=404, content={"success": False, "detail": "Patient not found"})
            
            person_uuid = person[0]
            rows_updated = 0
            
            if field_name == "sample_collection_date":
                # Find the exact sample by its current date and update it
                result = conn.execute(text("""
                    UPDATE laboratory_sample
                    SET date_sample_collected = CAST(:new_date AS DATE)
                    WHERE id IN (
                        SELECT ls.id
                        FROM laboratory_sample ls
                        INNER JOIN laboratory_test lt ON lt.id = ls.test_id
                        WHERE lt.lab_test_id = 16
                          AND ls.patient_uuid = :uuid
                          AND ls.archived = 0
                          AND CAST(ls.date_sample_collected AS DATE) = CAST(:sample_date AS DATE)
                        LIMIT 1
                    )
                """), {"new_date": new_value, "uuid": person_uuid, "sample_date": sample_date})
                rows_updated = result.rowcount
                
            elif field_name == "vl_result":
                # Find the result linked to this sample date and update it
                result = conn.execute(text("""
                    UPDATE laboratory_result
                    SET result_reported = :new_result
                    WHERE id IN (
                        SELECT lr.id
                        FROM laboratory_result lr
                        INNER JOIN laboratory_test lt ON lt.id = lr.test_id
                        INNER JOIN laboratory_sample ls ON ls.test_id = lt.id
                        WHERE lt.lab_test_id = 16
                          AND lr.patient_uuid = :uuid
                          AND lr.archived = 0
                          AND CAST(ls.date_sample_collected AS DATE) = CAST(:sample_date AS DATE)
                        ORDER BY ls.date_sample_collected DESC
                        LIMIT 1
                    )
                """), {"new_result": new_value, "uuid": person_uuid, "sample_date": sample_date})
                rows_updated = result.rowcount
                
            elif field_name == "result_date":
                # Find the result linked to this sample date and update its result date
                result = conn.execute(text("""
                    UPDATE laboratory_result
                    SET date_result_reported = CAST(:new_date AS DATE)
                    WHERE id IN (
                        SELECT lr.id
                        FROM laboratory_result lr
                        INNER JOIN laboratory_test lt ON lt.id = lr.test_id
                        INNER JOIN laboratory_sample ls ON ls.test_id = lt.id
                        WHERE lt.lab_test_id = 16
                          AND lr.patient_uuid = :uuid
                          AND lr.archived = 0
                          AND CAST(ls.date_sample_collected AS DATE) = CAST(:sample_date AS DATE)
                        ORDER BY ls.date_sample_collected DESC
                        LIMIT 1
                    )
                """), {"new_date": new_value, "uuid": person_uuid, "sample_date": sample_date})
                rows_updated = result.rowcount
            
            conn.commit()
            
            logger.info(f"✅ VL {field_name} updated: {rows_updated} row(s) affected")
            
            if rows_updated == 0:
                return JSONResponse(status_code=404, content={"success": False, "detail": f"No VL record found for sample date {sample_date}"})
            
            return {"success": True, "message": f"VL {field_name} updated successfully", "rows_updated": rows_updated}
            
    except Exception as e:
        logger.error(f"VL update error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})

@app.delete("/api/patients/delete-vl")
async def delete_viral_load(request: Request):
    """
    Delete an entire Viral Load record from EMR.
    Archives both laboratory_sample and laboratory_result by matching sample date.
    """
    try:
        from app.database import emr_engine
        from sqlalchemy import text
        
        body = await request.json()
        hospital_number = body.get("hospital_number")
        sample_date = body.get("sample_date")
        user = request.headers.get("X-User", "Unknown")
        
        if not hospital_number or not sample_date:
            return JSONResponse(status_code=400, content={"success": False, "detail": "Missing required fields"})
        
        logger.info(f"🗑️ Deleting VL record for {hospital_number} dated {sample_date} by {user}")
        
        with emr_engine.connect() as conn:
            # Get person_uuid
            person = conn.execute(text(
                "SELECT uuid FROM patient_person WHERE hospital_number = :hn AND archived = 0"
            ), {"hn": hospital_number}).fetchone()
            
            if not person:
                return JSONResponse(status_code=404, content={"success": False, "detail": "Patient not found"})
            
            person_uuid = person[0]
            
            # Find the sample ID first
            sample_record = conn.execute(text("""
                SELECT ls.id, ls.test_id
                FROM laboratory_sample ls
                INNER JOIN laboratory_test lt ON lt.id = ls.test_id
                WHERE lt.lab_test_id = 16
                  AND ls.patient_uuid = :uuid
                  AND ls.archived = 0
                  AND CAST(ls.date_sample_collected AS DATE) = CAST(:sample_date AS DATE)
                ORDER BY ls.date_sample_collected DESC
                LIMIT 1
            """), {"uuid": person_uuid, "sample_date": sample_date}).fetchone()
            
            if not sample_record:
                return JSONResponse(status_code=404, content={"success": False, "detail": f"No VL sample found for date {sample_date}"})
            
            sample_id = sample_record[0]
            test_id = sample_record[1]
            
            # Archive the sample
            conn.execute(text("""
                UPDATE laboratory_sample SET archived = 1
                WHERE id = :sample_id
            """), {"sample_id": sample_id})
            
            # Archive the result
            result = conn.execute(text("""
                UPDATE laboratory_result SET archived = 1
                WHERE test_id = :test_id
                  AND patient_uuid = :uuid
                  AND archived = 0
            """), {"test_id": test_id, "uuid": person_uuid})
            
            conn.commit()
            
            logger.info(f"✅ VL record deleted: sample_id={sample_id}, test_id={test_id}, results_archived={result.rowcount}")
            
            # Log the deletion
            try:
                from app.database import dqa_engine
                from app.models.dqa_models import CorrectionLog
                from sqlalchemy.orm import Session
                
                with Session(dqa_engine) as dqa_session:
                    log = CorrectionLog(
                        hospital_number=hospital_number,
                        person_uuid=person_uuid,
                        field_corrected="vl_record_deleted",
                        old_value=f"Sample Date: {sample_date}",
                        new_value="DELETED",
                        corrected_by=user,
                        record_type="viral_load"
                    )
                    dqa_session.add(log)
                    dqa_session.commit()
            except Exception as log_error:
                logger.warning(f"Correction log failed: {log_error}")
            
            return {"success": True, "message": f"VL record for {sample_date} deleted successfully"}
            
    except Exception as e:
        logger.error(f"VL delete error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"success": False, "detail": str(e)})
# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
        log_level="info"
    )